"""
consolidador_mensual.py — CONSOLIDADOR MENSUAL SAP (módulo separado, posterior
al V2 diario).

Toma SAP diarios YA VALIDADOS por el usuario (archivos `SAP_TIQ_DD-MM-YYYY.xlsx`
ya generados y publicados por el flujo diario) y produce UN NUEVO archivo SAP
GLOBAL mensual: `SAP_GLOBAL_TIQ_<MES>_<AÑO>.xlsx`.

MUY IMPORTANTE — este módulo es de solo lectura sobre sus entradas:

  - los SAP diarios de origen se abren EXCLUSIVAMENTE en modo lectura
    (`openpyxl.load_workbook(..., read_only=True)`); nunca se llama
    `.save()` sobre ellos;
  - la plantilla SAP maestra nunca se abre en modo escritura: se copia con
    `shutil.copyfile` y solo la COPIA (la ruta `--salida`) se abre para
    escribir;
  - no vuelve a interpretar reglas contables ni a revisar vouchers/CI/ATC/
    USD: solo comprueba que cada SAP diario ya generado es estructuralmente
    válido (cabecera, partidas, cuadre propio) y copia sus partidas TAL
    CUAL al SAP global, sin agrupar ni resumir;
  - no se conecta a Google Drive: opera exclusivamente sobre archivos ya
    materializados localmente (por Cowork, en una etapa posterior);
  - en el SAP global, la columna Cargo/Haber que no corresponde a una
    partida (DEBE con Haber, o HABER con Cargo) se escribe como celda
    REALMENTE VACÍA (`None`), nunca como 0/0.00 explícito: SAP no acepta
    ese cero en la columna opuesta (`_celda_importe_o_vacia`).

Estructura de la plantilla SAP (idéntica a la usada por sap_writer.py — hoja
EXACTA "1", cabecera fila 10, partidas desde fila 16, mismas columnas); ver
sap_writer.py para la referencia canónica del layout. Este módulo NO importa
sap_writer.py: implementa su propia escritura mínima para no acoplarse a la
forma del "asiento" del motor diario, que no aplica aquí.

Uso:
    python consolidador_mensual.py \\
        --anio 2026 \\
        --mes 8 \\
        --sap-dir /ruta/sap_diarios \\
        --plantilla /ruta/Plantilla_SAP_maestra.xlsx \\
        --salida /ruta/SAP_GLOBAL_TIQ_AGOSTO_2026.xlsx

    Opcionalmente, para consolidar solo una lista explícita de SAP ya
    aprobados (se omite el escaneo de --sap-dir):

    python consolidador_mensual.py \\
        --anio 2026 --mes 8 \\
        --archivos-lista /ruta/SAP_TIQ_01-08-2026.xlsx /ruta/SAP_TIQ_02-08-2026.xlsx \\
        --plantilla /ruta/Plantilla_SAP_maestra.xlsx \\
        --salida /ruta/SAP_GLOBAL_TIQ_AGOSTO_2026.xlsx
"""

import argparse
import calendar
import datetime
import hashlib
import json
import os
import re
import shutil
import sys
import warnings
from decimal import Decimal

import openpyxl

from excel_io import money_str, to_decimal


# ---------------------------------------------------------------------------
# Constantes de layout SAP (idénticas a sap_writer.py; definidas aquí de
# forma independiente para no importar ese módulo).
# ---------------------------------------------------------------------------

_HOJA_SAP = "1"
_FILA_CABECERA = 10
_FILA_PRIMERA_PARTIDA = 16
_FORMATO_FECHA_CORTA = "dd/mm/yyyy"

_CABECERA_ESPERADA = {"B": "BO01", "C": "DB", "H": "BOB", "L": "CAJA TIQUIPAYA"}

# Cabecera del SAP GLOBAL mensual (sección 8 de la especificación).
_SOCIEDAD_GLOBAL = "BO01"
_TIPO_ASIENTO_GLOBAL = "DB"
_MONEDA_GLOBAL = "BOB"
_REFERENCIA_GLOBAL = "CAJA TIQUIPAYA"

_MES_ABREV = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC",
}
_MES_NOMBRE = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

_RE_SAP_DIARIO = re.compile(r"^SAP_TIQ_(\d{2})-(\d{2})-(\d{4})\.xlsx$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Utilidades básicas
# ---------------------------------------------------------------------------

def _abrir_libro(ruta, **kwargs):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return openpyxl.load_workbook(ruta, **kwargs)


def _sha256_archivo(ruta):
    hasher = hashlib.sha256()
    with open(ruta, "rb") as f:
        for bloque in iter(lambda: f.read(1 << 16), b""):
            hasher.update(bloque)
    return hasher.hexdigest()


def _texto_celda(valor):
    if valor is None:
        return None
    return str(valor).strip()


def ultimo_dia_mes(anio, mes):
    ultimo = calendar.monthrange(anio, mes)[1]
    return datetime.date(anio, mes, ultimo)


def _es_temporal(nombre):
    return nombre.startswith("~$") or nombre.startswith(".") or nombre.lower().endswith(".tmp")


def _fecha_desde_nombre_sap(nombre):
    m = _RE_SAP_DIARIO.match(nombre)
    if not m:
        return None
    dia, mes, anio = m.groups()
    return datetime.date(int(anio), int(mes), int(dia))


# ---------------------------------------------------------------------------
# Sección 3 — selección de SAP diarios del mes solicitado
# ---------------------------------------------------------------------------

def seleccionar_sap_directorio(sap_dir, anio, mes):
    """Escanea `sap_dir` y devuelve, ordenadas cronológicamente, las rutas
    absolutas de los `SAP_TIQ_DD-MM-YYYY.xlsx` que correspondan a `anio`/
    `mes`. Ignora `SAP_GLOBAL_*` (no matchea el patrón), archivos
    temporales/ocultos y cualquier otro mes o nombre no compatible."""
    if not os.path.isdir(sap_dir):
        raise RuntimeError(f"SAP_DIR_NO_ENCONTRADO: {sap_dir}")

    candidatos = []
    for nombre in sorted(os.listdir(sap_dir)):
        if _es_temporal(nombre):
            continue
        fecha = _fecha_desde_nombre_sap(nombre)
        if fecha is None:
            continue
        if fecha.year != anio or fecha.month != mes:
            continue
        candidatos.append((fecha, os.path.abspath(os.path.join(sap_dir, nombre))))

    candidatos.sort(key=lambda t: t[0])
    return [ruta for _, ruta in candidatos]


def resolver_archivos(args):
    """Si se pasa --archivos-lista, se consolidan EXCLUSIVAMENTE esos
    archivos (sin filtrar por --sap-dir/año/mes), ordenados cronológicamente
    cuando el nombre permite determinar la fecha (si no, se conserva el
    orden recibido). Sin --archivos-lista, se usa el escaneo por defecto de
    --sap-dir (sección 3)."""
    if args.archivos_lista:
        rutas = [os.path.abspath(r) for r in args.archivos_lista]
        for ruta in rutas:
            if not os.path.isfile(ruta):
                raise RuntimeError(f"ARCHIVO_LISTA_NO_ENCONTRADO: {ruta}")

        indexadas = list(enumerate(rutas))

        def _clave(item):
            idx, ruta = item
            fecha = _fecha_desde_nombre_sap(os.path.basename(ruta))
            return (0, fecha) if fecha else (1, idx)

        indexadas.sort(key=_clave)
        return [ruta for _, ruta in indexadas]

    if not args.sap_dir:
        raise RuntimeError("FALTA_SAP_DIR_O_ARCHIVOS_LISTA")
    return seleccionar_sap_directorio(args.sap_dir, args.anio, args.mes)


# ---------------------------------------------------------------------------
# Sección 4 — duplicados (SHA256)
# ---------------------------------------------------------------------------

def detectar_duplicados(rutas):
    """Agrupa `rutas` por nombre de archivo y compara SHA256 dentro de cada
    grupo. Devuelve (rutas_unicas, sha256_por_archivo, duplicados_identicos,
    duplicados_diferentes):

      - rutas_unicas: una ruta por nombre (la primera, en el orden recibido),
        en el mismo orden cronológico de `rutas`;
      - sha256_por_archivo: {ruta_absoluta: sha256} de TODAS las rutas
        recibidas (trazabilidad completa, incluidas las ignoradas);
      - duplicados_identicos: [{"nombre", "rutas_ignoradas", "sha256"}, ...]
        — mismo nombre, mismo contenido byte a byte: se usa una sola copia;
      - duplicados_diferentes: [{"nombre", "rutas", "sha256"}, ...] — mismo
        nombre, contenido distinto: BLOQUEA la consolidación (nunca se
        elige arbitrariamente).
    """
    por_nombre = {}
    for ruta in rutas:
        por_nombre.setdefault(os.path.basename(ruta), []).append(ruta)

    sha256_por_archivo = {}
    rutas_unicas = []
    duplicados_identicos = []
    duplicados_diferentes = []

    for nombre, lista_rutas in por_nombre.items():
        hashes = {ruta: _sha256_archivo(ruta) for ruta in lista_rutas}
        sha256_por_archivo.update(hashes)
        valores_hash = set(hashes.values())

        if len(lista_rutas) == 1:
            rutas_unicas.append(lista_rutas[0])
        elif len(valores_hash) == 1:
            rutas_unicas.append(lista_rutas[0])
            duplicados_identicos.append({
                "nombre": nombre,
                "rutas_ignoradas": lista_rutas[1:],
                "sha256": next(iter(valores_hash)),
            })
        else:
            duplicados_diferentes.append({
                "nombre": nombre,
                "rutas": lista_rutas,
                "sha256": hashes,
            })

    orden = {ruta: i for i, ruta in enumerate(rutas)}
    rutas_unicas.sort(key=lambda r: orden[r])

    return rutas_unicas, sha256_por_archivo, duplicados_identicos, duplicados_diferentes


# ---------------------------------------------------------------------------
# Sección 2 — guardarraíles absolutos sobre --salida
# ---------------------------------------------------------------------------

def validar_guardarrieles_salida(ruta_salida, ruta_plantilla, rutas_origen, force):
    """Nunca permite que --salida coincida con la plantilla, con un SAP
    origen, o con el patrón de nombre de un SAP diario; y exige --force
    explícito para reemplazar una salida global ya existente. --force
    jamás habilita reemplazar un SAP diario ni la plantilla: solo se abre
    en modo escritura la propia ruta de --salida."""
    salida_abs = os.path.abspath(ruta_salida)

    if os.path.abspath(ruta_plantilla) == salida_abs:
        raise RuntimeError("RUTA_SALIDA_IGUAL_A_PLANTILLA")

    rutas_origen_abs = {os.path.abspath(r) for r in rutas_origen}
    if salida_abs in rutas_origen_abs:
        raise RuntimeError("RUTA_SALIDA_IGUAL_A_SAP_ORIGEN")

    nombre_salida = os.path.basename(salida_abs)
    if _RE_SAP_DIARIO.match(nombre_salida):
        raise RuntimeError(
            f"RUTA_SALIDA_NOMBRE_SAP_DIARIO: '{nombre_salida}' coincide con "
            f"el patrón de un SAP diario (SAP_TIQ_DD-MM-YYYY.xlsx); la "
            f"salida global nunca puede llevar ese nombre."
        )

    if os.path.exists(salida_abs) and not force:
        raise RuntimeError(
            f"SALIDA_YA_EXISTE_SIN_FORCE: '{salida_abs}' ya existe. Use "
            f"--force para reemplazarla (--force solo afecta la salida "
            f"GLOBAL, nunca un SAP diario ni la plantilla)."
        )


# ---------------------------------------------------------------------------
# Sección 5 — validación mínima de cada SAP diario (solo lectura)
# ---------------------------------------------------------------------------

def leer_y_validar_sap_diario(ruta):
    """Abre `ruta` EXCLUSIVAMENTE en modo lectura (read_only, data_only) y
    valida su estructura mínima: hoja "1", cabecera fija, y partidas desde
    la fila 16 (cuentas no vacías, Cargo/Haber consistentes por partida,
    Cargo total = Haber total del propio SAP). No reinterpreta reglas
    contables ni vuelve a revisar vouchers/CI/ATC/USD.

    Devuelve {"partidas": [...], "problemas": [...], "cargo_total": Decimal,
    "haber_total": Decimal}. `partidas` conserva, sin alterar, Sociedad,
    Cuenta, TextoPosicion, Cargo, Haber, CentroBeneficio, FechaValor,
    Asignacion y XREF1/2/3 tal cual están en el archivo.
    """
    try:
        wb = _abrir_libro(ruta, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — archivo ilegible, se reporta y bloquea
        return {
            "partidas": [], "problemas": [f"SAP_ILEGIBLE:{exc}"],
            "cargo_total": Decimal("0.00"), "haber_total": Decimal("0.00"),
        }

    problemas = []
    partidas = []
    cargo_total = Decimal("0.00")
    haber_total = Decimal("0.00")

    try:
        if _HOJA_SAP not in wb.sheetnames:
            return {
                "partidas": [], "problemas": ["HOJA_1_NO_ENCONTRADA"],
                "cargo_total": Decimal("0.00"), "haber_total": Decimal("0.00"),
            }
        ws = wb[_HOJA_SAP]

        for col, esperado in _CABECERA_ESPERADA.items():
            valor = _texto_celda(ws[f"{col}{_FILA_CABECERA}"].value)
            if valor != esperado:
                problemas.append(
                    f"CABECERA_{col}{_FILA_CABECERA}_ESPERADO_{esperado!r}_OBTENIDO_{valor!r}"
                )

        fila = _FILA_PRIMERA_PARTIDA
        while True:
            sociedad_val = ws[f"B{fila}"].value
            cuenta_val = ws[f"C{fila}"].value
            if sociedad_val in (None, "") and cuenta_val in (None, ""):
                break

            cuenta_txt = _texto_celda(cuenta_val)
            if not cuenta_txt:
                problemas.append(f"PARTIDA_{fila}_CUENTA_VACIA")

            cargo_val = ws[f"E{fila}"].value
            haber_val = ws[f"F{fila}"].value
            try:
                cargo_dec = to_decimal(cargo_val)
                haber_dec = to_decimal(haber_val)
            except ValueError:
                problemas.append(f"PARTIDA_{fila}_IMPORTE_NO_NUMERICO")
                cargo_dec = haber_dec = Decimal("0.00")

            if (cargo_dec > 0) == (haber_dec > 0):
                # ambos > 0, o ambos == 0: no es una partida Cargo/Haber válida
                problemas.append(f"PARTIDA_{fila}_CARGO_HABER_INCONSISTENTE")

            fecha_valor_val = ws[f"O{fila}"].value
            if fecha_valor_val is not None and not isinstance(
                fecha_valor_val, (datetime.date, datetime.datetime)
            ):
                problemas.append(f"PARTIDA_{fila}_FECHA_VALOR_INVALIDA")

            partidas.append({
                "sociedad": sociedad_val,
                "cuenta_mayor": cuenta_val,
                "texto_posicion": ws[f"D{fila}"].value,
                "cargo": cargo_dec,
                "haber": haber_dec,
                "centro_beneficio": ws[f"L{fila}"].value,
                "fecha_valor": fecha_valor_val,
                "asignacion": ws[f"R{fila}"].value,
                "xref1": ws[f"U{fila}"].value,
                "xref2": ws[f"V{fila}"].value,
                "xref3": ws[f"W{fila}"].value,
            })
            cargo_total += cargo_dec
            haber_total += haber_dec
            fila += 1

        if not partidas:
            problemas.append("SAP_SIN_PARTIDAS")
        elif cargo_total != haber_total:
            problemas.append(
                f"CARGO_TOTAL_DISTINTO_DE_HABER_TOTAL:cargo_{money_str(cargo_total)}"
                f"_haber_{money_str(haber_total)}"
            )
    finally:
        wb.close()

    return {
        "partidas": partidas,
        "problemas": problemas,
        "cargo_total": cargo_total,
        "haber_total": haber_total,
    }


# ---------------------------------------------------------------------------
# Sección 7-9 — plantilla global, cabecera y escritura detallada (sin
# agrupar) de todas las partidas.
# ---------------------------------------------------------------------------

def construir_metadata_cabecera_global(anio, mes):
    return {
        "sociedad": _SOCIEDAD_GLOBAL,
        "tipo_asiento": _TIPO_ASIENTO_GLOBAL,
        "fecha_registro": ultimo_dia_mes(anio, mes),
        "fecha_contabilizacion": ultimo_dia_mes(anio, mes),
        "mes": mes,
        "texto_cabecera": f"INGRESOS {_MES_ABREV[mes]} CBBA",
        "moneda": _MONEDA_GLOBAL,
        "referencia": _REFERENCIA_GLOBAL,
    }


def _celda_importe_o_vacia(importe):
    """SAP no acepta un 0/0.00 explícito en la columna Cargo/Haber que no
    corresponde a la partida (partida DEBE con Haber=0, o partida HABER
    con Cargo=0): esa celda debe quedar REALMENTE VACÍA. Devuelve el
    importe tal cual si es > 0, o None (celda vacía) en caso contrario.
    Nunca inventa ni ajusta el importe positivo real."""
    if importe is not None and importe > 0:
        return importe
    return None


def escribir_sap_global(partidas, ruta_plantilla, ruta_salida, metadata):
    """Copia `ruta_plantilla` a `ruta_salida` (nunca abre la plantilla en
    modo escritura) y escribe, sobre la copia, la cabecera global y TODAS
    las partidas recibidas, en el mismo orden, sin agrupar ni sumar."""
    shutil.copyfile(ruta_plantilla, ruta_salida)

    wb = _abrir_libro(ruta_salida)
    try:
        ws = wb[_HOJA_SAP]

        fila = _FILA_CABECERA
        ws[f"B{fila}"] = metadata["sociedad"]
        ws[f"C{fila}"] = metadata["tipo_asiento"]
        celda_reg = ws[f"D{fila}"]
        celda_reg.value = metadata["fecha_registro"]
        celda_reg.number_format = _FORMATO_FECHA_CORTA
        celda_cont = ws[f"E{fila}"]
        celda_cont.value = metadata["fecha_contabilizacion"]
        celda_cont.number_format = _FORMATO_FECHA_CORTA
        ws[f"F{fila}"] = metadata["mes"]
        ws[f"G{fila}"] = metadata["texto_cabecera"]
        ws[f"H{fila}"] = metadata["moneda"]
        ws[f"L{fila}"] = metadata["referencia"]

        for offset, p in enumerate(partidas):
            fila_p = _FILA_PRIMERA_PARTIDA + offset
            ws[f"B{fila_p}"] = p["sociedad"]
            ws[f"C{fila_p}"] = p["cuenta_mayor"]
            if p["texto_posicion"] is not None:
                ws[f"D{fila_p}"] = p["texto_posicion"]

            cargo_valor = _celda_importe_o_vacia(p["cargo"])
            haber_valor = _celda_importe_o_vacia(p["haber"])

            celda_cargo = ws[f"E{fila_p}"]
            celda_cargo.value = cargo_valor
            if cargo_valor is not None:
                celda_cargo.number_format = "0.00"

            celda_haber = ws[f"F{fila_p}"]
            celda_haber.value = haber_valor
            if haber_valor is not None:
                celda_haber.number_format = "0.00"

            ws[f"L{fila_p}"] = p["centro_beneficio"]
            if p["fecha_valor"] is not None:
                celda_fv = ws[f"O{fila_p}"]
                celda_fv.value = p["fecha_valor"]
                celda_fv.number_format = _FORMATO_FECHA_CORTA
            if p["asignacion"] is not None:
                ws[f"R{fila_p}"] = p["asignacion"]
            if p["xref1"] is not None:
                ws[f"U{fila_p}"] = p["xref1"]
            if p["xref2"] is not None:
                ws[f"V{fila_p}"] = p["xref2"]
            if p["xref3"] is not None:
                ws[f"W{fila_p}"] = p["xref3"]

        wb.save(ruta_salida)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Sección 11 — trazabilidad (RESULTADO_GLOBAL_TIQ_<MES>_<AÑO>.json)
# ---------------------------------------------------------------------------

def nombre_sap_global(anio, mes):
    return f"SAP_GLOBAL_TIQ_{_MES_NOMBRE[mes]}_{anio}.xlsx"


def nombre_resultado_json(anio, mes):
    return f"RESULTADO_GLOBAL_TIQ_{_MES_NOMBRE[mes]}_{anio}.json"


# ---------------------------------------------------------------------------
# Orquestador
# ---------------------------------------------------------------------------

def ejecutar_consolidacion(args):
    if not os.path.isfile(args.plantilla):
        raise RuntimeError(f"PLANTILLA_NO_ENCONTRADA: {args.plantilla}")

    rutas_candidatas = resolver_archivos(args)

    validar_guardarrieles_salida(args.salida, args.plantilla, rutas_candidatas, args.force)

    rutas_unicas, sha256_por_archivo, duplicados_identicos, duplicados_diferentes = \
        detectar_duplicados(rutas_candidatas)

    blockers = []
    for dup in duplicados_diferentes:
        blockers.append(f"DUPLICADO_SAP_DIFERENTE:{dup['nombre']}")

    partidas_por_archivo = {}
    if not duplicados_diferentes:
        for ruta in rutas_unicas:
            resultado_archivo = leer_y_validar_sap_diario(ruta)
            partidas_por_archivo[ruta] = resultado_archivo["partidas"]
            for problema in resultado_archivo["problemas"]:
                blockers.append(f"SAP_INVALIDO:{os.path.basename(ruta)}:{problema}")

    if not rutas_unicas and not duplicados_diferentes:
        blockers.append("SIN_SAP_PARA_CONSOLIDAR")

    todas_partidas = []
    cargo_global = haber_global = diferencia = None
    cantidad_partidas = None

    if not blockers:
        for ruta in rutas_unicas:
            todas_partidas.extend(partidas_por_archivo[ruta])

        cargo_global = sum((p["cargo"] for p in todas_partidas), Decimal("0.00"))
        haber_global = sum((p["haber"] for p in todas_partidas), Decimal("0.00"))
        diferencia = cargo_global - haber_global
        cantidad_partidas = len(todas_partidas)

        if diferencia != 0:
            blockers.append(
                f"CUADRE_GLOBAL_DESCUADRADO:cargo_{money_str(cargo_global)}"
                f"_haber_{money_str(haber_global)}_diferencia_{money_str(diferencia)}"
            )

    ruta_global_generado = None
    if not blockers:
        metadata = construir_metadata_cabecera_global(args.anio, args.mes)
        escribir_sap_global(todas_partidas, args.plantilla, args.salida, metadata)
        ruta_global_generado = os.path.abspath(args.salida)

    estado = "VALIDADO_PENDIENTE_PUBLICACION" if not blockers else "ERROR_REVISAR"

    resultado_json = {
        "anio": args.anio,
        "mes": args.mes,
        "mes_nombre": _MES_NOMBRE[args.mes],
        "fecha_generacion": datetime.datetime.now().isoformat(timespec="seconds"),
        "cantidad_sap_incluidos": len(rutas_unicas),
        "sap_incluidos": [os.path.basename(r) for r in rutas_unicas],
        "sha256_sap_origen": {
            os.path.basename(r): sha256_por_archivo[r] for r in rutas_unicas
        },
        "duplicados_identicos_ignorados": duplicados_identicos,
        "duplicados_diferentes_encontrados": duplicados_diferentes,
        "cantidad_partidas": cantidad_partidas,
        "cargo_global": money_str(cargo_global) if cargo_global is not None else None,
        "haber_global": money_str(haber_global) if haber_global is not None else None,
        "diferencia": money_str(diferencia) if diferencia is not None else None,
        "ruta_global_generado": ruta_global_generado,
        "blockers": blockers,
        "estado": estado,
    }

    ruta_json = os.path.join(
        os.path.dirname(os.path.abspath(args.salida)) or ".",
        nombre_resultado_json(args.anio, args.mes),
    )
    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(resultado_json, f, ensure_ascii=False, indent=2)

    resultado_json["ruta_resultado_json"] = ruta_json
    return resultado_json


def construir_parser():
    parser = argparse.ArgumentParser(
        description="Consolidador mensual SAP — Caja Tiquipaya V2. Toma SAP "
                     "diarios YA VALIDADOS (solo lectura) y genera un NUEVO "
                     "SAP GLOBAL mensual, sin agrupar ni resumir partidas."
    )
    parser.add_argument("--anio", required=True, type=int, help="Año del mes a consolidar (YYYY)")
    parser.add_argument("--mes", required=True, type=int, choices=range(1, 13), metavar="1-12",
                         help="Mes numérico a consolidar (1-12)")
    parser.add_argument("--sap-dir", default=None,
                         help="Directorio local con los SAP_TIQ_DD-MM-YYYY.xlsx (ignorado si se usa --archivos-lista)")
    parser.add_argument("--plantilla", required=True, help="Ruta local a la plantilla SAP maestra (nunca se modifica)")
    parser.add_argument("--salida", required=True, help="Ruta del NUEVO SAP_GLOBAL_TIQ_<MES>_<AÑO>.xlsx a generar")
    parser.add_argument("--archivos-lista", nargs="+", default=None,
                         help="Lista explícita de rutas de SAP diarios aprobados a consolidar "
                              "(si se pasa, se ignora --sap-dir y no se filtra por año/mes)")
    parser.add_argument("--force", action="store_true",
                         help="Permite reemplazar --salida si ya existe (nunca afecta SAP diarios ni la plantilla)")
    return parser


def main(argv=None):
    parser = construir_parser()
    args = parser.parse_args(argv)

    try:
        resultado = ejecutar_consolidacion(args)
    except (ValueError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    print(f"Consolidador mensual {args.mes:02d}/{args.anio}: estado={resultado['estado']}")
    print(f"SAP incluidos: {resultado['cantidad_sap_incluidos']}")
    if resultado["blockers"]:
        print("Blockers:")
        for b in resultado["blockers"]:
            print(f"  - {b}")
    if resultado["ruta_global_generado"]:
        print(f"SAP global generado: {resultado['ruta_global_generado']}")
    print(f"Resultado JSON: {resultado['ruta_resultado_json']}")

    return 0 if resultado["estado"] == "VALIDADO_PENDIENTE_PUBLICACION" else 1


if __name__ == "__main__":
    sys.exit(main())
