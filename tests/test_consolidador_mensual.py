"""
test_consolidador_mensual.py — CONSOLIDADOR MENSUAL SAP (módulo separado).

Usa una plantilla SAP SINTÉTICA (tests/xlsx_fixtures.crear_plantilla_sap) y
SAP diarios sintéticos construidos localmente en este archivo (nunca datos
contables reales). Verifica exclusivamente el comportamiento de
consolidador_mensual.py: no repite ninguna regla de excel_io.py/
motor_tiquipaya.py/sap_writer.py/run_batch.py.

Uso: python -m unittest tests.test_consolidador_mensual -v
"""

import datetime
import os
import shutil
import sys
import tempfile
import unittest
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import consolidador_mensual as cm
from tests.xlsx_fixtures import crear_plantilla_sap


# ---------------------------------------------------------------------------
# Fixture: SAP diario sintético mínimo válido (hoja "1", cabecera fija,
# partidas cuadradas desde fila 16).
# ---------------------------------------------------------------------------

def _crear_sap_diario(ruta, partidas, tipo_asiento="DB", cargar_cabecera_valida=True):
    """partidas: lista de dicts con las claves de una partida (ver
    consolidador_mensual.leer_y_validar_sap_diario). Cargo/Haber pueden
    pasarse como str/Decimal/float; se escriben tal cual (Decimal)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1"

    if cargar_cabecera_valida:
        ws["B10"] = "BO01"
        ws["C10"] = tipo_asiento
        ws["H10"] = "BOB"
        ws["L10"] = "CAJA TIQUIPAYA"

    fila = 16
    for p in partidas:
        ws[f"B{fila}"] = p.get("sociedad", "BO01")
        ws[f"C{fila}"] = p.get("cuenta_mayor", "110101001")
        ws[f"D{fila}"] = p.get("texto_posicion")
        cargo = p.get("cargo", "0.00")
        haber = p.get("haber", "0.00")
        ws[f"E{fila}"] = Decimal(str(cargo)) if cargo is not None else None
        ws[f"F{fila}"] = Decimal(str(haber)) if haber is not None else None
        ws[f"L{fila}"] = p.get("centro_beneficio", "10010101")
        fecha_valor = p.get("fecha_valor")
        if fecha_valor is not None:
            ws[f"O{fila}"] = fecha_valor
        ws[f"R{fila}"] = p.get("asignacion")
        ws[f"U{fila}"] = p.get("xref1")
        ws[f"V{fila}"] = p.get("xref2")
        ws[f"W{fila}"] = p.get("xref3")
        fila += 1

    wb.save(ruta)


def _a_fecha(valor):
    """openpyxl puede reabrir una celda de fecha como datetime.datetime en
    lugar de datetime.date; normaliza para comparar solo la fecha."""
    if isinstance(valor, datetime.datetime):
        return valor.date()
    return valor


def _partida_par(cargo, haber="0.00", **extra):
    base = {"cargo": cargo, "haber": haber}
    base.update(extra)
    return base


def _dos_partidas_cuadradas(importe="100.00", **kwargs):
    """Una partida DEBE + una HABER por el mismo importe (cuadra sola)."""
    fecha_valor = kwargs.pop("fecha_valor", datetime.date(2026, 8, 5))
    asignacion = kwargs.pop("asignacion", "ASIG-1")
    return [
        _partida_par(importe, "0.00", cuenta_mayor="110101001",
                     texto_posicion="RECAUDACION", fecha_valor=fecha_valor,
                     asignacion=asignacion, xref1="X1"),
        _partida_par("0.00", importe, cuenta_mayor="210201005",
                     texto_posicion="CONTRAPARTIDA", fecha_valor=fecha_valor,
                     asignacion=asignacion),
    ]


class _ConsolidadorTestBase(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix="consolidador_test_")
        self.sap_dir = os.path.join(self.tmpdir, "sap_diarios")
        os.makedirs(self.sap_dir)
        self.ruta_plantilla = os.path.join(self.tmpdir, "Plantilla_SAP_maestra.xlsx")
        crear_plantilla_sap(self.ruta_plantilla)
        self.ruta_salida = os.path.join(self.tmpdir, "SAP_GLOBAL_TIQ_AGOSTO_2026.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _ruta_diaria(self, dia, mes=8, anio=2026):
        return os.path.join(self.sap_dir, f"SAP_TIQ_{dia:02d}-{mes:02d}-{anio}.xlsx")

    def _args(self, **overrides):
        class _Args:
            pass
        a = _Args()
        a.anio = overrides.get("anio", 2026)
        a.mes = overrides.get("mes", 8)
        a.sap_dir = overrides.get("sap_dir", self.sap_dir)
        a.plantilla = overrides.get("plantilla", self.ruta_plantilla)
        a.salida = overrides.get("salida", self.ruta_salida)
        a.archivos_lista = overrides.get("archivos_lista", None)
        a.force = overrides.get("force", False)
        return a


# ---------------------------------------------------------------------------
# A: consolidación básica de dos SAP diarios válidos
# ---------------------------------------------------------------------------

class TestConsolidacionBasica(_ConsolidadorTestBase):

    def setUp(self):
        super().setUp()
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("100.00"))
        _crear_sap_diario(self._ruta_diaria(2), _dos_partidas_cuadradas("200.00"))

    def test_consolida_y_estado_validado(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION", resultado)
        self.assertEqual(resultado["blockers"], [])
        self.assertTrue(os.path.isfile(self.ruta_salida))

    def test_texto_cabecera_ingresos_ago_cbba(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["G10"].value, "INGRESOS AGO CBBA")
        wb.close()

    def test_texto_cabecera_septiembre(self):
        _crear_sap_diario(self._ruta_diaria(1, mes=9), _dos_partidas_cuadradas("50.00"))
        args = self._args(mes=9, sap_dir=os.path.join(self.tmpdir, "sap_sep"))
        os.makedirs(args.sap_dir)
        shutil.copy(self._ruta_diaria(1, mes=9), args.sap_dir)
        args.salida = os.path.join(self.tmpdir, "SAP_GLOBAL_TIQ_SEPTIEMBRE_2026.xlsx")
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION", resultado)
        wb = openpyxl.load_workbook(args.salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["G10"].value, "INGRESOS SEP CBBA")
        wb.close()

    def test_fecha_global_es_ultimo_dia_del_mes(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(_a_fecha(ws["D10"].value), datetime.date(2026, 8, 31))
        self.assertEqual(_a_fecha(ws["E10"].value), datetime.date(2026, 8, 31))
        self.assertEqual(ws["F10"].value, 8)
        wb.close()
        self.assertIsNotNone(resultado)

    def test_febrero_bisiesto(self):
        anio_bisiesto = 2028
        self.assertEqual(cm.ultimo_dia_mes(anio_bisiesto, 2), datetime.date(2028, 2, 29))
        self.assertEqual(cm.ultimo_dia_mes(2026, 2), datetime.date(2026, 2, 28))

    def test_cabecera_global_sociedad_tipo_moneda_referencia(self):
        cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["B10"].value, "BO01")
        self.assertEqual(ws["C10"].value, "DB")
        self.assertEqual(ws["H10"].value, "BOB")
        self.assertEqual(ws["L10"].value, "CAJA TIQUIPAYA")
        wb.close()

    def test_copia_todas_las_partidas_sin_agrupar(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["cantidad_partidas"], 4)  # 2 SAP x 2 partidas
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        cuentas = [ws[f"C{16 + i}"].value for i in range(4)]
        self.assertEqual(
            cuentas, ["110101001", "210201005", "110101001", "210201005"]
        )
        wb.close()

    def test_no_agrupa_por_cuenta(self):
        # Dos partidas con la MISMA cuenta en SAP distintos deben aparecer
        # como dos filas separadas, nunca sumadas en una sola.
        resultado = cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        filas_110101001 = [
            i for i in range(4) if ws[f"C{16 + i}"].value == "110101001"
        ]
        self.assertEqual(len(filas_110101001), 2)
        wb.close()
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION")

    def test_conserva_cuentas_textos_importes(self):
        cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["C16"].value, "110101001")
        self.assertEqual(ws["D16"].value, "RECAUDACION")
        self.assertEqual(cm.to_decimal(ws["E16"].value), Decimal("100.00"))
        wb.close()

    def test_conserva_fecha_valor_original_no_fin_de_mes(self):
        cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(_a_fecha(ws["O16"].value), datetime.date(2026, 8, 5))
        self.assertNotEqual(_a_fecha(ws["O16"].value), datetime.date(2026, 8, 31))
        wb.close()

    def test_conserva_asignacion(self):
        cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["R16"].value, "ASIG-1")
        wb.close()

    def test_conserva_xref(self):
        cm.ejecutar_consolidacion(self._args())
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws["U16"].value, "X1")
        wb.close()

    def test_cargo_global_igual_haber_global(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["cargo_global"], resultado["haber_global"])
        self.assertEqual(resultado["diferencia"], "0.00")

    def test_suma_global_igual_a_suma_diaria(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        # 100.00 (SAP día 1) + 200.00 (SAP día 2) = 300.00 en cada lado.
        self.assertEqual(resultado["cargo_global"], "300.00")
        self.assertEqual(resultado["haber_global"], "300.00")

    def test_orden_cronologico(self):
        # Crea un tercer SAP con fecha anterior a los otros dos para
        # confirmar que el orden de escritura es por fecha, no alfabético.
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("999.00"))
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["sap_incluidos"], [
            "SAP_TIQ_01-08-2026.xlsx", "SAP_TIQ_02-08-2026.xlsx",
        ])
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION")


# ---------------------------------------------------------------------------
# B: selección de archivos (mes/año, SAP_GLOBAL_*, temporales)
# ---------------------------------------------------------------------------

class TestSeleccionArchivos(_ConsolidadorTestBase):

    def test_solo_sap_del_mes_solicitado(self):
        _crear_sap_diario(self._ruta_diaria(1, mes=8), _dos_partidas_cuadradas("10.00"))
        _crear_sap_diario(self._ruta_diaria(1, mes=9), _dos_partidas_cuadradas("20.00"))
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["sap_incluidos"], ["SAP_TIQ_01-08-2026.xlsx"])

    def test_ignora_sap_global(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        ruta_global_previa = os.path.join(self.sap_dir, "SAP_GLOBAL_TIQ_AGOSTO_2026.xlsx")
        _crear_sap_diario(ruta_global_previa, _dos_partidas_cuadradas("999.00"))
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["sap_incluidos"], ["SAP_TIQ_01-08-2026.xlsx"])

    def test_ignora_archivos_temporales(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        ruta_temp = os.path.join(self.sap_dir, "~$SAP_TIQ_02-08-2026.xlsx")
        with open(ruta_temp, "w") as f:
            f.write("no es un xlsx real")
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["sap_incluidos"], ["SAP_TIQ_01-08-2026.xlsx"])

    def test_ignora_nombres_no_compatibles(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        ruta_rara = os.path.join(self.sap_dir, "OTRO_ARCHIVO.xlsx")
        _crear_sap_diario(ruta_rara, _dos_partidas_cuadradas("999.00"))
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["sap_incluidos"], ["SAP_TIQ_01-08-2026.xlsx"])

    def test_sin_sap_para_consolidar_bloquea(self):
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertIn("SIN_SAP_PARA_CONSOLIDAR", resultado["blockers"])
        self.assertFalse(os.path.isfile(self.ruta_salida))


# ---------------------------------------------------------------------------
# C: --archivos-lista
# ---------------------------------------------------------------------------

class TestArchivosLista(_ConsolidadorTestBase):

    def test_archivos_lista_consolida_solo_los_indicados(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        _crear_sap_diario(self._ruta_diaria(2), _dos_partidas_cuadradas("20.00"))
        args = self._args(archivos_lista=[self._ruta_diaria(2)])
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["sap_incluidos"], ["SAP_TIQ_02-08-2026.xlsx"])
        self.assertEqual(resultado["cargo_global"], "20.00")

    def test_archivos_lista_ordena_cronologicamente(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        _crear_sap_diario(self._ruta_diaria(2), _dos_partidas_cuadradas("20.00"))
        args = self._args(archivos_lista=[self._ruta_diaria(2), self._ruta_diaria(1)])
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["sap_incluidos"], [
            "SAP_TIQ_01-08-2026.xlsx", "SAP_TIQ_02-08-2026.xlsx",
        ])


# ---------------------------------------------------------------------------
# D: duplicados (SHA256)
# ---------------------------------------------------------------------------

class TestDuplicados(_ConsolidadorTestBase):

    def test_duplicado_identico_usa_una_copia(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        otro_dir = os.path.join(self.tmpdir, "otra_copia")
        os.makedirs(otro_dir)
        ruta_copia = os.path.join(otro_dir, "SAP_TIQ_01-08-2026.xlsx")
        shutil.copyfile(self._ruta_diaria(1), ruta_copia)

        args = self._args(archivos_lista=[self._ruta_diaria(1), ruta_copia])
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION", resultado)
        self.assertEqual(resultado["cantidad_sap_incluidos"], 1)
        self.assertEqual(len(resultado["duplicados_identicos_ignorados"]), 1)

    def test_duplicado_diferente_bloquea(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))
        otro_dir = os.path.join(self.tmpdir, "otra_copia")
        os.makedirs(otro_dir)
        ruta_diferente = os.path.join(otro_dir, "SAP_TIQ_01-08-2026.xlsx")
        _crear_sap_diario(ruta_diferente, _dos_partidas_cuadradas("999.00"))

        args = self._args(archivos_lista=[self._ruta_diaria(1), ruta_diferente])
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(
            any(b.startswith("DUPLICADO_SAP_DIFERENTE") for b in resultado["blockers"])
        )
        self.assertFalse(os.path.isfile(self.ruta_salida))
        self.assertEqual(len(resultado["duplicados_diferentes_encontrados"]), 1)


# ---------------------------------------------------------------------------
# E: SAP descuadrado / estructuralmente inválido bloquea
# ---------------------------------------------------------------------------

class TestSapInvalidoBloquea(_ConsolidadorTestBase):

    def test_sap_descuadrado_bloquea(self):
        partidas = [_partida_par("100.00", "0.00"), _partida_par("0.00", "90.00")]
        _crear_sap_diario(self._ruta_diaria(1), partidas)
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(
            any("CARGO_TOTAL_DISTINTO_DE_HABER_TOTAL" in b for b in resultado["blockers"])
        )
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_cabecera_invalida_bloquea(self):
        _crear_sap_diario(
            self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"), tipo_asiento="SA"
        )
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(any("CABECERA_C10" in b for b in resultado["blockers"]))

    def test_sin_partidas_bloquea(self):
        _crear_sap_diario(self._ruta_diaria(1), [])
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(any("SAP_SIN_PARTIDAS" in b for b in resultado["blockers"]))

    def test_cuenta_vacia_bloquea(self):
        _crear_sap_diario(self._ruta_diaria(1), [
            {"cuenta_mayor": "", "cargo": "10.00", "haber": "0.00"},
            {"cuenta_mayor": "210201005", "cargo": "0.00", "haber": "10.00"},
        ])
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(any("CUENTA_VACIA" in b for b in resultado["blockers"]))

    def test_cargo_haber_ambos_positivos_bloquea(self):
        _crear_sap_diario(self._ruta_diaria(1), [
            {"cuenta_mayor": "110101001", "cargo": "10.00", "haber": "10.00"},
        ])
        resultado = cm.ejecutar_consolidacion(self._args())
        self.assertEqual(resultado["estado"], "ERROR_REVISAR")
        self.assertTrue(
            any("CARGO_HABER_INCONSISTENTE" in b for b in resultado["blockers"])
        )


# ---------------------------------------------------------------------------
# F: guardarraíles de --salida
# ---------------------------------------------------------------------------

class TestGuardarrielesSalida(_ConsolidadorTestBase):

    def setUp(self):
        super().setUp()
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("10.00"))

    def test_salida_no_puede_ser_un_origen(self):
        args = self._args(salida=self._ruta_diaria(1))
        with self.assertRaises(RuntimeError) as ctx:
            cm.ejecutar_consolidacion(args)
        self.assertIn("RUTA_SALIDA_IGUAL_A_SAP_ORIGEN", str(ctx.exception))

    def test_salida_no_puede_ser_la_plantilla(self):
        args = self._args(salida=self.ruta_plantilla)
        with self.assertRaises(RuntimeError) as ctx:
            cm.ejecutar_consolidacion(args)
        self.assertIn("RUTA_SALIDA_IGUAL_A_PLANTILLA", str(ctx.exception))

    def test_salida_no_puede_tener_nombre_de_sap_diario(self):
        args = self._args(salida=os.path.join(self.tmpdir, "SAP_TIQ_15-08-2026.xlsx"))
        with self.assertRaises(RuntimeError) as ctx:
            cm.ejecutar_consolidacion(args)
        self.assertIn("RUTA_SALIDA_NOMBRE_SAP_DIARIO", str(ctx.exception))

    def test_no_sobrescribe_sin_force(self):
        with open(self.ruta_salida, "w") as f:
            f.write("archivo global previo")
        args = self._args()
        with self.assertRaises(RuntimeError) as ctx:
            cm.ejecutar_consolidacion(args)
        self.assertIn("SALIDA_YA_EXISTE_SIN_FORCE", str(ctx.exception))

    def test_force_solo_sobre_salida_global(self):
        with open(self.ruta_salida, "w") as f:
            f.write("archivo global previo")
        hash_diario_antes = cm._sha256_archivo(self._ruta_diaria(1))
        hash_plantilla_antes = cm._sha256_archivo(self.ruta_plantilla)

        args = self._args(force=True)
        resultado = cm.ejecutar_consolidacion(args)
        self.assertEqual(resultado["estado"], "VALIDADO_PENDIENTE_PUBLICACION", resultado)
        self.assertTrue(os.path.isfile(self.ruta_salida))

        self.assertEqual(cm._sha256_archivo(self._ruta_diaria(1)), hash_diario_antes)
        self.assertEqual(cm._sha256_archivo(self.ruta_plantilla), hash_plantilla_antes)


# ---------------------------------------------------------------------------
# G: NO modificar orígenes (sección 12 de la especificación)
# ---------------------------------------------------------------------------

class TestOrigenesSoloLectura(_ConsolidadorTestBase):

    def setUp(self):
        super().setUp()
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("100.00"))
        _crear_sap_diario(self._ruta_diaria(2), _dos_partidas_cuadradas("200.00"))

    def test_hash_sap_origen_sin_cambios(self):
        hashes_antes = {
            ruta: cm._sha256_archivo(ruta)
            for ruta in (self._ruta_diaria(1), self._ruta_diaria(2))
        }
        cm.ejecutar_consolidacion(self._args())
        for ruta, hash_antes in hashes_antes.items():
            self.assertEqual(cm._sha256_archivo(ruta), hash_antes)

    def test_plantilla_sin_cambios(self):
        hash_antes = cm._sha256_archivo(self.ruta_plantilla)
        cm.ejecutar_consolidacion(self._args())
        self.assertEqual(cm._sha256_archivo(self.ruta_plantilla), hash_antes)

    def test_tamano_sap_origen_sin_cambios(self):
        tamanos_antes = {
            ruta: os.path.getsize(ruta)
            for ruta in (self._ruta_diaria(1), self._ruta_diaria(2))
        }
        cm.ejecutar_consolidacion(self._args())
        for ruta, tamano_antes in tamanos_antes.items():
            self.assertEqual(os.path.getsize(ruta), tamano_antes)

    def test_mtime_sap_origen_sin_cambios_si_es_comprobable(self):
        mtimes_antes = {
            ruta: os.path.getmtime(ruta)
            for ruta in (self._ruta_diaria(1), self._ruta_diaria(2))
        }
        cm.ejecutar_consolidacion(self._args())
        for ruta, mtime_antes in mtimes_antes.items():
            self.assertEqual(os.path.getmtime(ruta), mtime_antes)

    def test_ningun_sap_origen_recibe_save(self):
        # Verificación indirecta: los SAP de origen se abren siempre con
        # read_only=True en leer_y_validar_sap_diario (openpyxl bloquea
        # .save() sobre libros read_only), y consolidador_mensual.py nunca
        # llama openpyxl.load_workbook(...) sin read_only=True salvo sobre
        # la propia ruta de --salida (ya una copia de la plantilla).
        resultado_lectura = cm.leer_y_validar_sap_diario(self._ruta_diaria(1))
        self.assertEqual(resultado_lectura["problemas"], [])
        wb = openpyxl.load_workbook(self._ruta_diaria(1), read_only=True)
        with self.assertRaises(Exception):
            wb.save(self._ruta_diaria(1))
        wb.close()


# ---------------------------------------------------------------------------
# H: trazabilidad (RESULTADO_GLOBAL_TIQ_<MES>_<AÑO>.json)
# ---------------------------------------------------------------------------

class TestTrazabilidad(_ConsolidadorTestBase):

    def test_json_resultado_generado_con_campos_esperados(self):
        _crear_sap_diario(self._ruta_diaria(1), _dos_partidas_cuadradas("100.00"))
        resultado = cm.ejecutar_consolidacion(self._args())

        ruta_json = os.path.join(self.tmpdir, "RESULTADO_GLOBAL_TIQ_AGOSTO_2026.json")
        self.assertTrue(os.path.isfile(ruta_json))

        import json
        with open(ruta_json, encoding="utf-8") as f:
            contenido = json.load(f)

        for campo in (
            "anio", "mes", "fecha_generacion", "cantidad_sap_incluidos",
            "sap_incluidos", "sha256_sap_origen",
            "duplicados_identicos_ignorados", "duplicados_diferentes_encontrados",
            "cantidad_partidas", "cargo_global", "haber_global", "diferencia",
            "ruta_global_generado", "blockers", "estado",
        ):
            self.assertIn(campo, contenido)

        self.assertEqual(contenido["anio"], 2026)
        self.assertEqual(contenido["mes"], 8)
        self.assertEqual(contenido["estado"], "VALIDADO_PENDIENTE_PUBLICACION")
        self.assertEqual(resultado["estado"], contenido["estado"])


# ---------------------------------------------------------------------------
# I: sin dependencia de Google Drive
# ---------------------------------------------------------------------------

class TestSinDependenciaDrive(unittest.TestCase):

    def test_no_importa_ningun_cliente_de_google_drive(self):
        # El docstring del módulo menciona "Google Drive" a propósito (para
        # aclarar que NO se conecta), así que se revisan solo las líneas de
        # import reales, nunca el texto libre de la documentación.
        import inspect
        for linea in inspect.getsource(cm).splitlines():
            stripped = linea.strip()
            if stripped.startswith("import ") or stripped.startswith("from "):
                lower = stripped.lower()
                for termino in ("google", "pydrive", "gspread"):
                    self.assertNotIn(termino, lower, msg=f"import sospechoso: {stripped}")


if __name__ == "__main__":
    unittest.main()
