"""
test_texto_posicion_fecha_valor.py — CORRECCIÓN FINAL ETAPA 6: transporte
de texto_posicion (SGTXT/columna D) y fecha_valor (VALUT/columna O) para
CI y VOUCHERS.

Reglas cubiertas (sin tocar ninguna regla contable: importes, cuentas,
asignaciones, cuadre, ATC, ALQUILERES, USD, cabecera/estructura SAP):

- CI: texto_posicion = columna "GLOSA ASIENTO COMUNICACIONES INTERNAS"
  literal (nunca se reconstruye). fecha_valor = columna "FECHA2" (nunca
  la fecha del cierre).
- VOUCHER: fecha_valor = columna "FECHA DE DEPOSITO" propia de cada
  depósito (nunca la fecha bancaria de MACROS). texto_posicion =
  "DEPOSITO BNB DD/MM/YYYY" con esa misma fecha.

Ejercita el parseo real de openpyxl (fixtures sintéticos, sin datos
contables reales), igual que test_regresion_sintetica.py, y también el
nivel de sap_writer.py (columnas D y O).

Uso: python -m unittest tests.test_texto_posicion_fecha_valor -v
"""

import os
import sys
import shutil
import tempfile
import unittest
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import motor_tiquipaya as motor
import sap_writer as sap
from tests import xlsx_fixtures as fx

FECHA_CIERRE = "2026-08-19"
NOMBRE_CIERRE = f"CIERRE {FECHA_CIERRE[8:10]}-{FECHA_CIERRE[5:7]}-{FECHA_CIERRE[0:4]}.xlsm"


def _sfc101():
    return {
        "total_movimiento": "325.00",
        "cobros_atc": "0.00",
        "dolares": "0.00",
        "depositos": [
            {"importe": "100.00", "fecha": "2026-08-19", "asignacion": "VCHA001", "banco": "BNB"},
            {"importe": "150.00", "fecha": "2026-08-20", "asignacion": "VCHA002", "banco": "BNB"},
        ],
        "ci": [
            {"total": "75.00", "cuenta": "210201005", "asignacion": "CI9001", "banco": "BNB",
             "glosa": "PAGO SERVICIOS AGUA", "fecha2": "2026-08-18"},
        ],
    }


def _sfc102():
    return {
        "total_movimiento": "0.00", "cobros_atc": "0.00", "dolares": "0.00",
        "depositos": [], "ci": [],
    }


def _macros_filas():
    return [
        (FECHA_CIERRE, "VCHA001", "100.00"),
        (FECHA_CIERRE, "VCHA002", "150.00"),
    ]


class _EjecutarBase(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.ruta_cierre = os.path.join(self._tmp.name, NOMBRE_CIERRE)
        self.ruta_macros = os.path.join(self._tmp.name, "MACROS.xlsm")
        self.ruta_atc = os.path.join(self._tmp.name, "ATC.xlsx")

        fx.crear_cierre(self.ruta_cierre, _sfc101(), _sfc102())
        fx.crear_macros(self.ruta_macros, _macros_filas())
        fx.crear_atc(self.ruta_atc, [])  # ATC bruto = 0.00: ATC_NO_APLICA, sin filas

        self.resultado_v2 = motor.ejecutar_v2(self.ruta_cierre, self.ruta_macros, self.ruta_atc)
        self.assertEqual(self.resultado_v2["estado"], "OK", self.resultado_v2)
        self.asiento = motor.construir_asiento(self.resultado_v2)
        self.assertEqual(self.asiento["estado"], "OK", self.asiento)

    def _partida(self, origen, asignacion=None):
        candidatas = [p for p in self.asiento["partidas"] if p["origen"] == origen]
        if asignacion is not None:
            candidatas = [p for p in candidatas if p["asignacion"] == asignacion]
        self.assertEqual(len(candidatas), 1, candidatas)
        return candidatas[0]


# ---------------------------------------------------------------------------
# 1-2, 6-8. CI: glosa literal en texto_posicion, FECHA2 en fecha_valor,
# importe/cuenta/asignación intactos.
# ---------------------------------------------------------------------------

class TestComunicacionesInternas(_EjecutarBase):

    def test_glosa_llega_literal_a_texto_posicion(self):
        ci = self._partida("CI", asignacion="CI9001")
        self.assertEqual(ci["texto_posicion"], "PAGO SERVICIOS AGUA")

    def test_fecha2_llega_a_fecha_valor(self):
        ci = self._partida("CI", asignacion="CI9001")
        self.assertEqual(ci["fecha_valor"], "2026-08-18")
        # nunca la fecha del cierre como reemplazo
        self.assertNotEqual(ci["fecha_valor"], FECHA_CIERRE)

    def test_importe_no_cambia(self):
        ci = self._partida("CI", asignacion="CI9001")
        self.assertEqual(ci["cargo"], "75.00")

    def test_cuenta_no_cambia(self):
        ci = self._partida("CI", asignacion="CI9001")
        self.assertEqual(ci["cuenta_mayor"], "210201005")

    def test_asignacion_no_cambia(self):
        ci = self._partida("CI", asignacion="CI9001")
        self.assertEqual(ci["asignacion"], "CI9001")


# ---------------------------------------------------------------------------
# 3-5, 9. VOUCHER: "DEPOSITO BNB DD/MM/YYYY", fecha propia por depósito,
# importe/cuenta/asignación intactos.
# ---------------------------------------------------------------------------

class TestVouchers(_EjecutarBase):

    def test_texto_posicion_deposito_bnb(self):
        v1 = self._partida("VOUCHER", asignacion="VCHA001")
        self.assertEqual(v1["texto_posicion"], "DEPOSITO BNB 19/08/2026")

    def test_fecha_valor_es_fecha_deposito(self):
        v1 = self._partida("VOUCHER", asignacion="VCHA001")
        self.assertEqual(v1["fecha_valor"], "2026-08-19")

    def test_dos_vouchers_fechas_distintas_texto_y_fecha_propios(self):
        v1 = self._partida("VOUCHER", asignacion="VCHA001")
        v2 = self._partida("VOUCHER", asignacion="VCHA002")
        self.assertEqual(v1["fecha_valor"], "2026-08-19")
        self.assertEqual(v1["texto_posicion"], "DEPOSITO BNB 19/08/2026")
        self.assertEqual(v2["fecha_valor"], "2026-08-20")
        self.assertEqual(v2["texto_posicion"], "DEPOSITO BNB 20/08/2026")

    def test_importe_cuenta_asignacion_no_cambian(self):
        v1 = self._partida("VOUCHER", asignacion="VCHA001")
        v2 = self._partida("VOUCHER", asignacion="VCHA002")
        self.assertEqual(v1["cargo"], "100.00")
        self.assertEqual(v1["cuenta_mayor"], "110103012")
        self.assertEqual(v2["cargo"], "150.00")
        self.assertEqual(v2["cuenta_mayor"], "110103012")


# ---------------------------------------------------------------------------
# 10. ATC permanece sin cambios (ATC_NO_APLICA con bruto=0.00, ya
# cubierto extensamente por tests/test_atc_preconciliado.py y
# tests/test_cruces.py — aquí solo se reconfirma dentro de este mismo
# escenario end-to-end).
# ---------------------------------------------------------------------------

class TestAtcSinCambios(_EjecutarBase):
    def test_atc_no_aplica_sin_lineas(self):
        origenes = {p["origen"] for p in self.asiento["partidas"]}
        self.assertNotIn("ATC_NETO", origenes)
        self.assertNotIn("ATC_COMISION", origenes)
        self.assertFalse(self.resultado_v2["detalle"]["atc_aplica"])
        self.assertEqual(self.resultado_v2["detalle"]["atc_estado"], "ATC_NO_APLICA")


# ---------------------------------------------------------------------------
# 11-12. SAP: columna D (texto_posicion) y columna O (fecha_valor)
# escritas correctamente para CI y VOUCHER.
# ---------------------------------------------------------------------------

class TestSapColumnasDyO(_EjecutarBase):
    def setUp(self):
        super().setUp()
        self.tmpdir_sap = tempfile.mkdtemp(prefix="sap_texto_fecha_")
        self.ruta_plantilla = os.path.join(self.tmpdir_sap, "plantilla.xlsx")
        self.ruta_salida = os.path.join(self.tmpdir_sap, "salida.xlsx")
        fx.crear_plantilla_sap(self.ruta_plantilla)
        self.metadata = {
            "tipo_asiento": "SA", "fecha_registro": FECHA_CIERRE,
            "fecha_contabilizacion": FECHA_CIERRE, "mes": "08",
            "texto_cabecera": "CAJA TIQUIPAYA 19-08-2026", "referencia": "TIQ-19082026",
        }

    def tearDown(self):
        shutil.rmtree(self.tmpdir_sap, ignore_errors=True)

    def test_sap_escribe_texto_posicion_y_fecha_valor(self):
        resumen = sap.generar_y_validar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        self.assertEqual(resumen["estado_sap"], "OK", resumen)

        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        for offset, p in enumerate(self.asiento["partidas"]):
            fila = 16 + offset
            self.assertEqual(ws[f"D{fila}"].value, p.get("texto_posicion"))
            self.assertEqual(ws[f"O{fila}"].value, p.get("fecha_valor"))
        wb.close()

        ci = self._partida("CI", asignacion="CI9001")
        idx_ci = self.asiento["partidas"].index(ci)
        v1 = self._partida("VOUCHER", asignacion="VCHA001")
        idx_v1 = self.asiento["partidas"].index(v1)

        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws[f"D{16 + idx_ci}"].value, "PAGO SERVICIOS AGUA")
        self.assertEqual(ws[f"O{16 + idx_ci}"].value, "2026-08-18")
        self.assertEqual(ws[f"D{16 + idx_v1}"].value, "DEPOSITO BNB 19/08/2026")
        self.assertEqual(ws[f"O{16 + idx_v1}"].value, "2026-08-19")
        wb.close()


if __name__ == "__main__":
    unittest.main()
