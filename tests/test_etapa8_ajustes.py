"""
test_etapa8_ajustes.py — ETAPA 8: ajustes finales SAP + control operativo.

Cubre, SIN reinterpretar contabilidad ni tocar ninguna regla ya validada
en ETAPAS 1-7 (importes, cuentas, asignaciones, cuadre, ATC preconciliado,
autocorrección 0/O, POSIBLE_TYPO, ALQUILERES, USD, estructura SAP, XREF):

  1-2. texto_posicion (SGTXT) autorizado para las 2 líneas HABER normales
       (UNIVERSO_SFC101/UNIVERSO_SFC102).
  3-6. prioridad de la fecha real de origen (CI: FECHA2: VOUCHER: FECHA
       DE DEPOSITO) sobre el fallback fecha_valor=fecha_cierre; HABER SFC
       y ATC sin fecha propia usan ese fallback.
  7-8. ninguna partida válida del asiento queda con fecha_valor vacía.
  9-10. en SAP, fecha_valor (columna O) se escribe como fecha Excel real
        (no texto), con number_format dd/mm/yyyy.
  11-15. cabecera SAP (D10/E10/F10) se deriva determinísticamente de la
         fecha real del cierre extraída por el motor (nunca fin de mes).
  16-18. importes, cuentas y asignaciones no cambian por ninguno de estos
         ajustes.

Usa un asiento construido a mano (mismo patrón que
tests/test_sap.py::_resultado_v2_ok), sin datos contables reales, para
poder fijar con precisión qué partidas traen fecha real y cuáles no.

Uso: python -m unittest tests.test_etapa8_ajustes -v
"""

import datetime
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import motor_tiquipaya as motor
import pipeline_tiquipaya as pipeline
import sap_writer as sap
from tests import xlsx_fixtures as fx


# ---------------------------------------------------------------------------
# Fixture: asiento con los 4 orígenes de partida (HABER x2, VOUCHER, CI,
# ATC_NETO, ATC_COMISION), parametrizado en las fechas propias de origen
# para poder probar la prioridad "fecha real > fallback fecha cierre".
# ---------------------------------------------------------------------------

def _resultado_v2_asiento_completo(fecha_cierre="2026-08-21", fecha_ci=None,
                                    fecha_deposito=None, fecha_bancaria_atc=None):
    detalle = {
        "sfc101_total": "1000.00",
        "sfc102_total": "500.00",
        "sfc101_haber": "1000.00",
        "sfc102_haber": "500.00",
        "alquileres_sfc101": "0.00",
        "alquileres_sfc102": "0.00",
        "vouchers_confirmados": [
            {
                "sfc": "SFC101",
                "importe": "300.00",
                "codigo_confirmado": "VCH0001",
                "codigo_informado": "VCH0001",
                "fecha_bancaria": "2026-08-22",  # nunca se usa para SAP
                "fecha_deposito": fecha_deposito,
                "estado": "MATCH_EXACTO",
            },
        ],
        "ci_validas": [
            {
                "sfc": "SFC102",
                "referencia": "FAC-0001",
                "importe": "100.00",
                "cuenta_contable": "210201005",
                "asignacion": "CI0001",
                "glosa": "PAGO PROVEEDOR AGUA",
                "fecha_ci": fecha_ci,
            },
        ],
        "atc_neto": {
            "importe": "1000.00",
            "codigo_confirmado": "ATC-CONF",
            "fecha_bancaria": fecha_bancaria_atc,
        },
        "atc_comision": {"importe": "100.00"},
        "atc_aplica": True,
        "dolares": "0.00",
    }
    return {
        "fecha": fecha_cierre,
        "universo_original": "1500.00",
        "alquileres": "0.00",
        "universo_ajustado": "1500.00",
        "componentes": {
            "vouchers": "300.00", "ci_operativas": "100.00",
            "atc_bruto": "1100.00", "dolares": "0.00",
        },
        "recaudacion_explicada": "1500.00",
        "diferencia": "0.00",
        "excepciones_bloqueantes": 0,
        "estado": "OK",
        "detalle": detalle,
    }


def _asiento(**kwargs):
    resultado = _resultado_v2_asiento_completo(**kwargs)
    asiento = motor.construir_asiento(resultado)
    return resultado, asiento


def _partida(asiento, origen):
    candidatas = [p for p in asiento["partidas"] if p["origen"] == origen]
    assert len(candidatas) == 1, candidatas
    return candidatas[0]


# ---------------------------------------------------------------------------
# 1-2. texto_posicion autorizado para HABER SFC101/SFC102.
# ---------------------------------------------------------------------------

class TestTextoPosicionHaberSfc(unittest.TestCase):
    def setUp(self):
        _, self.asiento = _asiento(fecha_cierre="2026-08-21")
        self.assertEqual(self.asiento["estado"], "OK", self.asiento)

    def test_sfc101_texto_recaudacion_caja_sfc101(self):
        sfc101 = _partida(self.asiento, "UNIVERSO_SFC101")
        self.assertEqual(sfc101["texto_posicion"], "RECAUDACION CAJA SFC101")

    def test_sfc102_texto_recaudacion_caja_sfc102(self):
        sfc102 = _partida(self.asiento, "UNIVERSO_SFC102")
        self.assertEqual(sfc102["texto_posicion"], "RECAUDACION CAJA SFC102")


# ---------------------------------------------------------------------------
# 3-6. Prioridad fecha real > fallback fecha cierre.
# ---------------------------------------------------------------------------

class TestPrioridadFechaRealSobreFallback(unittest.TestCase):

    def test_ci_conserva_fecha2_propia(self):
        _, asiento = _asiento(fecha_cierre="2026-08-21", fecha_ci="2026-08-15")
        ci = _partida(asiento, "CI")
        self.assertEqual(ci["fecha_valor"], "2026-08-15")
        self.assertNotEqual(ci["fecha_valor"], "2026-08-21")

    def test_voucher_conserva_fecha_de_deposito_propia(self):
        _, asiento = _asiento(fecha_cierre="2026-08-21", fecha_deposito="2026-08-10")
        voucher = _partida(asiento, "VOUCHER")
        self.assertEqual(voucher["fecha_valor"], "2026-08-10")
        self.assertNotEqual(voucher["fecha_valor"], "2026-08-21")

    def test_fecha_real_ci_y_voucher_nunca_se_reemplaza_por_fallback(self):
        _, asiento = _asiento(
            fecha_cierre="2026-08-21", fecha_ci="2026-08-15", fecha_deposito="2026-08-10",
        )
        ci = _partida(asiento, "CI")
        voucher = _partida(asiento, "VOUCHER")
        self.assertEqual(ci["fecha_valor"], "2026-08-15")
        self.assertEqual(voucher["fecha_valor"], "2026-08-10")

    def test_atc_sin_fecha_real_usa_fecha_cierre(self):
        _, asiento = _asiento(fecha_cierre="2026-08-21", fecha_bancaria_atc=None)
        atc_neto = _partida(asiento, "ATC_NETO")
        atc_comision = _partida(asiento, "ATC_COMISION")
        self.assertEqual(atc_neto["fecha_valor"], "2026-08-21")
        self.assertEqual(atc_comision["fecha_valor"], "2026-08-21")

    def test_haber_sfc_usa_fecha_cierre(self):
        _, asiento = _asiento(fecha_cierre="2026-08-21")
        sfc101 = _partida(asiento, "UNIVERSO_SFC101")
        sfc102 = _partida(asiento, "UNIVERSO_SFC102")
        self.assertEqual(sfc101["fecha_valor"], "2026-08-21")
        self.assertEqual(sfc102["fecha_valor"], "2026-08-21")


# ---------------------------------------------------------------------------
# 7-8. Ninguna partida válida queda con fecha_valor vacía.
# ---------------------------------------------------------------------------

class TestNingunaPartidaSinFechaValor(unittest.TestCase):
    def test_todas_las_partidas_tienen_fecha_valor(self):
        _, asiento = _asiento(fecha_cierre="2026-08-21")
        self.assertEqual(asiento["estado"], "OK")
        self.assertTrue(asiento["partidas"])
        for p in asiento["partidas"]:
            self.assertTrue(p.get("fecha_valor"), p)


# ---------------------------------------------------------------------------
# 16-18. Importes, cuentas y asignaciones no cambian por estos ajustes.
# ---------------------------------------------------------------------------

class TestSinImpactoEnValoresContables(unittest.TestCase):
    def setUp(self):
        _, self.asiento = _asiento(
            fecha_cierre="2026-08-21", fecha_ci="2026-08-15", fecha_deposito="2026-08-10",
        )

    def test_importes_no_cambian(self):
        sfc101 = _partida(self.asiento, "UNIVERSO_SFC101")
        sfc102 = _partida(self.asiento, "UNIVERSO_SFC102")
        voucher = _partida(self.asiento, "VOUCHER")
        ci = _partida(self.asiento, "CI")
        self.assertEqual(sfc101["haber"], "1000.00")
        self.assertEqual(sfc102["haber"], "500.00")
        self.assertEqual(voucher["cargo"], "300.00")
        self.assertEqual(ci["cargo"], "100.00")
        self.assertEqual(self.asiento["diferencia"], "0.00")

    def test_cuentas_no_cambian(self):
        sfc101 = _partida(self.asiento, "UNIVERSO_SFC101")
        voucher = _partida(self.asiento, "VOUCHER")
        ci = _partida(self.asiento, "CI")
        self.assertEqual(sfc101["cuenta_mayor"], "110101001")
        self.assertEqual(voucher["cuenta_mayor"], "110103012")
        self.assertEqual(ci["cuenta_mayor"], "210201005")

    def test_asignaciones_no_cambian(self):
        sfc101 = _partida(self.asiento, "UNIVERSO_SFC101")
        voucher = _partida(self.asiento, "VOUCHER")
        ci = _partida(self.asiento, "CI")
        self.assertEqual(sfc101["asignacion"], "SFC101")
        self.assertEqual(voucher["asignacion"], "VCH0001")
        self.assertEqual(ci["asignacion"], "CI0001")


# ---------------------------------------------------------------------------
# 9-10. SAP: fecha_valor (columna O) como fecha Excel real, dd/mm/yyyy.
# ---------------------------------------------------------------------------

class TestSapFechaValorTipoFechaExcel(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix="sap_etapa8_")
        self.addCleanup(shutil.rmtree, self.tmpdir, ignore_errors=True)
        self.ruta_plantilla = os.path.join(self.tmpdir, "plantilla.xlsx")
        self.ruta_salida = os.path.join(self.tmpdir, "salida.xlsx")
        fx.crear_plantilla_sap(self.ruta_plantilla)

        _, self.asiento = _asiento(
            fecha_cierre="2026-08-21", fecha_ci="2026-08-15", fecha_deposito="2026-08-10",
        )
        self.assertEqual(self.asiento["estado"], "OK")

        self.metadata = {
            "tipo_asiento": "SA", "fecha_registro": "2026-08-21",
            "fecha_contabilizacion": "2026-08-21", "mes": "08",
            "texto_cabecera": "CAJA TIQUIPAYA 21-08-2026", "referencia": "TIQ-21082026",
        }

    def test_fecha_valor_es_tipo_fecha_excel_y_sobrevive_reapertura(self):
        resumen = sap.generar_y_validar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        self.assertEqual(resumen["estado_sap"], "OK", resumen)

        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        for offset, p in enumerate(self.asiento["partidas"]):
            fila = 16 + offset
            valor_celda = ws[f"O{fila}"].value
            self.assertIsInstance(valor_celda, (datetime.date, datetime.datetime), p)
            self.assertEqual(sap._fecha_a_iso(valor_celda), p["fecha_valor"])
        wb.close()

    def test_columna_o_formato_corto_dd_mm_yyyy(self):
        sap.generar_y_validar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        for offset in range(len(self.asiento["partidas"])):
            fila = 16 + offset
            self.assertIn("dd/mm/yyyy", ws[f"O{fila}"].number_format.lower())
        wb.close()


# ---------------------------------------------------------------------------
# 11-15. Cabecera SAP derivada determinísticamente de la fecha del cierre.
# ---------------------------------------------------------------------------

class TestCabeceraDerivadaDelCierre(unittest.TestCase):
    def _generar(self, fecha_cierre):
        tmpdir = tempfile.mkdtemp(prefix="sap_cabecera_etapa8_")
        self.addCleanup(shutil.rmtree, tmpdir, ignore_errors=True)
        ruta_plantilla = os.path.join(tmpdir, "plantilla.xlsx")
        ruta_salida = os.path.join(tmpdir, "salida.xlsx")
        fx.crear_plantilla_sap(ruta_plantilla)

        _, asiento = _asiento(fecha_cierre=fecha_cierre)
        self.assertEqual(asiento["estado"], "OK")

        metadata = {
            "tipo_asiento": "SA", "texto_cabecera": "CAJA TIQUIPAYA", "referencia": "TIQ-TEST",
        }
        # ETAPA 8: fecha_registro/fecha_contabilizacion/mes NUNCA se
        # inventan a mano: se derivan de la fecha real del cierre con la
        # misma función que usa pipeline_tiquipaya.procesar_cierre_completo.
        metadata.update(pipeline.derivar_cabecera_fecha_cierre(fecha_cierre))

        resumen = sap.generar_y_validar_sap(asiento, ruta_plantilla, ruta_salida, metadata)
        self.assertEqual(resumen["estado_sap"], "OK", resumen)

        wb = openpyxl.load_workbook(ruta_salida, data_only=True)
        return wb, wb["1"]

    def test_cierre_21_08_produce_cabecera_21_08_no_31_08(self):
        wb, ws = self._generar("2026-08-21")
        self.assertEqual(sap._fecha_a_iso(ws["D10"].value), "2026-08-21")
        self.assertEqual(sap._fecha_a_iso(ws["E10"].value), "2026-08-21")
        self.assertNotEqual(sap._fecha_a_iso(ws["D10"].value), "2026-08-31")
        self.assertEqual(ws["F10"].value, "08")
        wb.close()

    def test_cierre_19_08_produce_cabecera_19_08(self):
        wb, ws = self._generar("2026-08-19")
        self.assertEqual(sap._fecha_a_iso(ws["D10"].value), "2026-08-19")
        self.assertEqual(sap._fecha_a_iso(ws["E10"].value), "2026-08-19")
        self.assertEqual(ws["F10"].value, "08")
        wb.close()

    def test_d10_e10_formato_corto_dd_mm_yyyy(self):
        wb, ws = self._generar("2026-08-21")
        self.assertIn("dd/mm/yyyy", ws["D10"].number_format.lower())
        self.assertIn("dd/mm/yyyy", ws["E10"].number_format.lower())
        wb.close()


if __name__ == "__main__":
    unittest.main()
