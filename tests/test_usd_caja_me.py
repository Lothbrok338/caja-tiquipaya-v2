"""
test_usd_caja_me.py — CIERRE DEFINITIVO: regla USD/DOLARES (Caja M/E).

Regla validada end-to-end por Cowork sobre el cierre real 05-08-2026 (ver
HANDOFF_CODE_V2.md): USD > 0.00 genera una única partida DEBE, cuenta
110101010 "Caja M/E", texto_posicion "RECAUDACION DOLARES", fecha_valor
igual a la fecha del cierre, asignación vacía (no existe fuente
autorizada), Sociedad BO01, Centro Beneficio 10010101. USD NO se concilia
contra banco/MACROS. USD > 0 ya NO produce "USD_CUENTA_PENDIENTE".

No reinterpreta ninguna otra regla: vouchers, CI, ATC, alquileres,
SFC101/SFC102, reglas SAP y control inmutable SHA256 quedan intactos.
Fixtures sintéticos (tests/xlsx_fixtures.py), sin datos contables reales.

Uso: python -m unittest tests.test_usd_caja_me -v
"""

import os
import shutil
import sys
import tempfile
import unittest
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import excel_io as io
import motor_tiquipaya as motor
import sap_writer as sap
from tests import xlsx_fixtures as fx


FECHA_CIERRE = "2026-08-05"
NOMBRE_CIERRE = f"CIERRE {FECHA_CIERRE[8:10]}-{FECHA_CIERRE[5:7]}-{FECHA_CIERRE[0:4]}.xlsm"

_DOLARES = "3606.00"


def _sfc101_con_usd():
    return {
        "total_movimiento": io.money_str(Decimal("225.00") + Decimal(_DOLARES)),
        "cobros_atc": "0.00",
        "dolares": _DOLARES,
        "depositos": [
            {"importe": "100.00", "fecha": FECHA_CIERRE, "asignacion": "VCHA001", "banco": "BNB"},
        ],
        "ci": [
            {"total": "125.00", "cuenta": "210201005", "asignacion": "CI9001", "banco": "BNB"},
        ],
    }


def _sfc102():
    return {
        "total_movimiento": "0.00", "cobros_atc": "0.00", "dolares": "0.00",
        "depositos": [], "ci": [],
    }


def _macros_filas():
    return [(FECHA_CIERRE, "VCHA001", "100.00")]


class _EjecutarBase(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.ruta_cierre = os.path.join(self._tmp.name, NOMBRE_CIERRE)
        self.ruta_macros = os.path.join(self._tmp.name, "MACROS.xlsm")
        self.ruta_atc = os.path.join(self._tmp.name, "ATC.xlsx")

        fx.crear_cierre(self.ruta_cierre, self._sfc101(), _sfc102())
        fx.crear_macros(self.ruta_macros, _macros_filas())
        fx.crear_atc(self.ruta_atc, [])  # ATC bruto = 0.00: ATC_NO_APLICA

        self.resultado_v2 = motor.ejecutar_v2(self.ruta_cierre, self.ruta_macros, self.ruta_atc)
        self.asiento = motor.construir_asiento(self.resultado_v2)

    def _sfc101(self):
        return _sfc101_con_usd()


# ---------------------------------------------------------------------------
# USD > 0: nunca USD_CUENTA_PENDIENTE, genera la partida DEBE 110101010.
# ---------------------------------------------------------------------------

class TestUsdEndToEnd(_EjecutarBase):

    def test_estado_ok_nunca_usd_cuenta_pendiente(self):
        self.assertEqual(self.resultado_v2["estado"], "OK", self.resultado_v2)
        self.assertNotEqual(self.resultado_v2["estado"], "USD_CUENTA_PENDIENTE")
        self.assertEqual(self.resultado_v2["diferencia"], "0.00")
        self.assertEqual(self.asiento["estado"], "OK", self.asiento)
        self.assertNotEqual(self.asiento["estado"], "USD_CUENTA_PENDIENTE")

    def test_usd_genera_partida_debe_110101010(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertEqual(usd["cuenta_mayor"], "110101010")
        self.assertEqual(usd["haber"], "0.00")

    def test_usd_importe_preservado(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertEqual(usd["cargo"], _DOLARES)

    def test_usd_texto_posicion_recaudacion_dolares(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertEqual(usd["texto_posicion"], "RECAUDACION DOLARES")

    def test_usd_fecha_valor_es_fecha_cierre(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertEqual(usd["fecha_valor"], FECHA_CIERRE)

    def test_usd_asignacion_vacia(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertIsNone(usd["asignacion"])

    def test_usd_sociedad_y_centro_beneficio(self):
        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        self.assertEqual(usd["sociedad"], "BO01")
        self.assertEqual(usd["centro_beneficio"], "10010101")

    def test_otras_reglas_intactas(self):
        origenes = {p["origen"] for p in self.asiento["partidas"]}
        self.assertIn("UNIVERSO_SFC101", origenes)
        self.assertIn("UNIVERSO_SFC102", origenes)
        self.assertIn("VOUCHER", origenes)
        self.assertIn("CI", origenes)
        self.assertNotIn("ALQUILERES", origenes)
        # ATC_NO_APLICA en este fixture (bruto=0.00): sin líneas ATC.
        self.assertNotIn("ATC_NETO", origenes)
        self.assertNotIn("ATC_COMISION", origenes)

    def test_cuadre_no_se_altera(self):
        self.assertEqual(self.asiento["total_cargo"], self.asiento["total_haber"])
        self.assertEqual(self.asiento["diferencia"], "0.00")


# ---------------------------------------------------------------------------
# USD = 0: no genera partida, y el estado nunca fue USD_CUENTA_PENDIENTE.
# ---------------------------------------------------------------------------

class TestUsdCeroNoGeneraPartida(_EjecutarBase):
    def _sfc101(self):
        datos = _sfc101_con_usd()
        datos["dolares"] = "0.00"
        datos["total_movimiento"] = "225.00"
        return datos

    def test_usd_cero_no_genera_partida_dolares(self):
        self.assertEqual(self.asiento["estado"], "OK", self.asiento)
        origenes = {p["origen"] for p in self.asiento["partidas"]}
        self.assertNotIn("DOLARES", origenes)

    def test_estado_nunca_usd_cuenta_pendiente(self):
        self.assertNotEqual(self.resultado_v2["estado"], "USD_CUENTA_PENDIENTE")
        self.assertNotEqual(self.asiento["estado"], "USD_CUENTA_PENDIENTE")


# ---------------------------------------------------------------------------
# SAP: la partida USD se escribe correctamente (cuenta/texto/importe/
# centro beneficio/fecha valor/asignación vacía).
# ---------------------------------------------------------------------------

class TestUsdEnSap(_EjecutarBase):
    def setUp(self):
        super().setUp()
        self.assertEqual(self.asiento["estado"], "OK", self.asiento)

        self.tmpdir_sap = tempfile.mkdtemp(prefix="sap_usd_")
        self.addCleanup(shutil.rmtree, self.tmpdir_sap, ignore_errors=True)
        self.ruta_plantilla = os.path.join(self.tmpdir_sap, "plantilla.xlsx")
        self.ruta_salida = os.path.join(self.tmpdir_sap, "salida.xlsx")
        fx.crear_plantilla_sap(self.ruta_plantilla)
        self.metadata = {
            "tipo_asiento": "SA", "fecha_registro": FECHA_CIERRE,
            "fecha_contabilizacion": FECHA_CIERRE, "mes": "08",
            "texto_cabecera": "CAJA TIQUIPAYA 05-08-2026", "referencia": "TIQ-05082026",
        }

    def test_sap_escribe_partida_usd_correctamente(self):
        resumen = sap.generar_y_validar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        self.assertEqual(resumen["estado_sap"], "OK", resumen)

        usd = next(p for p in self.asiento["partidas"] if p["origen"] == "DOLARES")
        idx = self.asiento["partidas"].index(usd)
        fila = 16 + idx

        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws[f"C{fila}"].value, "110101010")
        self.assertEqual(ws[f"D{fila}"].value, "RECAUDACION DOLARES")
        self.assertEqual(sap._decimal_celda(ws[f"E{fila}"].value), Decimal(_DOLARES))
        self.assertEqual(sap._decimal_celda(ws[f"F{fila}"].value), Decimal("0.00"))
        self.assertEqual(ws[f"L{fila}"].value, "10010101")
        self.assertEqual(sap._fecha_a_iso(ws[f"O{fila}"].value), FECHA_CIERRE)
        self.assertIsNone(ws[f"R{fila}"].value)
        wb.close()


if __name__ == "__main__":
    unittest.main()
