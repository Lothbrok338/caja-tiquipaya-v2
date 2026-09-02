"""xlsx_fixtures.py — constructores mínimos de archivos .xlsx/.xlsm SINTÉTICOS
para las pruebas de excel_io.py y motor_tiquipaya.py.

Ningún dato contable real: todos los importes, códigos y cuentas son
inventados exclusivamente para pruebas unitarias/de integración. Reproduce
solo la estructura mínima de hojas y columnas que excel_io.py exige, para
poder probar el parseo real (openpyxl) sin subir Excel reales al repo.
"""

import openpyxl


def crear_cierre(ruta, sfc101, sfc102):
    """sfc101 / sfc102: dict con:
        total_movimiento, cobros_atc, dolares, total_ci (opcional)
        depositos: [{"deposito": "DEPOSITO", "importe":, "fecha":,
                      "asignacion":, "banco":}, ...] o {"separador": True}
                     para una fila vacía intermedia.
        ci: [{"n":, "factura":, "total":, "cuenta":, "asignacion":,
               "banco":, "fecha": opcional}]
    """
    wb = openpyxl.Workbook()
    ws101 = wb.active
    ws101.title = "SFC101"
    _llenar_resumen(ws101, sfc101)

    ws102 = wb.create_sheet("SFC102")
    _llenar_resumen(ws102, sfc102)

    ws_ci101 = wb.create_sheet("COMUNICACIONES INTERNAS SFC101")
    _llenar_ci(ws_ci101, sfc101.get("ci", []))

    ws_ci102 = wb.create_sheet("COMUNICACIONES INTERNAS SFC102")
    _llenar_ci(ws_ci102, sfc102.get("ci", []))

    wb.save(ruta)


def _llenar_resumen(ws, datos):
    ws.append(["TOTAL MOVIMIENTO DEL DIA", datos["total_movimiento"]])
    ws.append(["COBROS ATC", datos["cobros_atc"]])
    ws.append(["TOTAL COMUNICACIONES INTERNAS", datos.get("total_ci", "0.00")])
    ws.append(["DOLARES", datos.get("dolares", "0.00")])
    ws.append([None, None, None, None, None])
    # Encabezado de COMPOSICIÓN DE DEPÓSITOS: las columnas IMPORTE/FECHA/
    # ASIGNACION/BANCO deben ir en la MISMA fila que la etiqueta (así lo
    # exige excel_io._leer_composicion_depositos).
    ws.append(["COMPOSICION DE DEPOSITOS", "IMPORTE Bs", "FECHA DE DEPOSITO",
                "ASIGNACION", "BANCO"])
    for dep in datos.get("depositos", []):
        if dep.get("separador"):
            ws.append([None, None, None, None, None])
            continue
        ws.append([
            dep.get("deposito", "DEPOSITO"), dep["importe"], dep.get("fecha"),
            dep.get("asignacion"), dep.get("banco", "BNB"),
        ])


def _llenar_ci(ws, filas):
    con_fecha = any("fecha" in f for f in filas)
    header = ["N°", "N° DE FACTURA", "TOTAL C.I.", "CUENTA CONTABLE BANCO",
               "ASIGNACION", "BANCO"]
    if con_fecha:
        header.append("FECHA")
    ws.append(header)
    for i, f in enumerate(filas, start=1):
        row = [
            f.get("n", i), f.get("factura", f"FAC-{i}"), f["total"],
            f.get("cuenta"), f.get("asignacion"), f.get("banco", "BNB"),
        ]
        if con_fecha:
            row.append(f.get("fecha"))
        ws.append(row)


def crear_macros(ruta, filas, header_repetido_en=None,
                  hoja="Tablas Dinamicas Profesional"):
    """filas: [(fecha, codigo, credito), ...].
    header_repetido_en: índice de fila (0-based sobre `filas`) donde
    insertar el encabezado repetido, para probar la deduplicación."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    header = ["Fecha", "Código de Asignación", "Créditos"]
    ws.append(header)
    for i, (fecha, codigo, credito) in enumerate(filas):
        if header_repetido_en is not None and i == header_repetido_en:
            ws.append(header)
        ws.append([fecha, codigo, credito])
    wb.save(ruta)


def crear_atc(ruta, filas):
    """filas: [(fecha, tipo, monto), ...]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["FECHA", "TIPO", "CUENTA CONTABLE", "DETALLE", "MONTO",
                "CÓDIGO ASIENTO", "ASIGNACION"])
    for fecha, tipo, monto in filas:
        ws.append([fecha, tipo, None, None, monto, None, None])
    wb.save(ruta)


def _llenar_atc_preconciliado(ws, filas):
    """filas: [(fecha, tipo, cuenta_contable, detalle, monto, asignacion), ...]
    ASIGNACION puede ser "REVISAR" para probar esa regla."""
    ws.append(["FECHA", "TIPO", "CUENTA CONTABLE", "DETALLE", "MONTO", "ASIGNACION"])
    for fila in filas:
        ws.append(list(fila))


def crear_atc_preconciliado(ruta, filas, hoja="ATC TIQUIPAYA"):
    """ATC TIQUIPAYA (ETAPA 6) como archivo separado, mismo esquema de
    columnas que dentro del maestro único. Sirve para probar el lector
    en aislamiento (excel_io.leer_atc_mensual) y para el "flujo anterior
    con ATC separado" ahora en formato preconciliado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    _llenar_atc_preconciliado(ws, filas)
    wb.save(ruta)


def crear_maestro_unico(ruta, macros_filas, atc_filas, header_repetido_en=None,
                         hoja_macros="Tablas Dinamicas Profesional",
                         hoja_atc="ATC TIQUIPAYA"):
    """MAESTRO MENSUAL ÚNICO (ETAPA 6): un solo workbook con las dos hojas
    relevantes. macros_filas: igual formato que crear_macros(). atc_filas:
    igual formato que crear_atc_preconciliado()."""
    wb = openpyxl.Workbook()
    ws_macros = wb.active
    ws_macros.title = hoja_macros
    header_macros = ["Fecha", "Código de Asignación", "Créditos"]
    ws_macros.append(header_macros)
    for i, (fecha, codigo, credito) in enumerate(macros_filas):
        if header_repetido_en is not None and i == header_repetido_en:
            ws_macros.append(header_macros)
        ws_macros.append([fecha, codigo, credito])

    ws_atc = wb.create_sheet(hoja_atc)
    _llenar_atc_preconciliado(ws_atc, atc_filas)

    wb.save(ruta)


def crear_plantilla_sap(ruta, hoja="1"):
    """Plantilla SAP sintética mínima (ETAPA 6), sin datos contables reales.

    Reproduce solo la estructura exigida por sap_writer.py: hoja EXACTA
    "1", cabecera en fila 10 (columnas B/C/D/E/F/G/H/L) y partidas desde
    fila 16 (columnas B/C/D/E/F/L/O/R/U/V/W), todo vacío para que
    sap_writer.py lo complete. Incluye contenido de ejemplo en columnas
    NO autorizadas (A, I, K, N, P, Q, S, T, X) para poder probar que
    sap_writer.py nunca las toca.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja

    ws["A9"] = "PLANTILLA SAP SINTETICA DE PRUEBA"
    ws["A10"] = "FILA CABECERA"
    ws["I10"] = "NO TOCAR"
    ws["K10"] = "NO TOCAR"
    ws["N10"] = "NO TOCAR"

    ws["A16"] = "FILA PRIMERA PARTIDA"
    ws["I16"] = "NO TOCAR"
    ws["K16"] = "NO TOCAR"
    ws["N16"] = "NO TOCAR"
    ws["P16"] = "NO TOCAR"
    ws["Q16"] = "NO TOCAR"
    ws["S16"] = "NO TOCAR"
    ws["T16"] = "NO TOCAR"
    ws["X16"] = "NO TOCAR"

    wb.save(ruta)
