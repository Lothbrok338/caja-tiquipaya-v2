"""
test_sap.py — ETAPA 6: pruebas de sap_writer.py (generación y validación
determinística del archivo SAP).

Usa una plantilla SAP SINTÉTICA (tests/xlsx_fixtures.crear_plantilla_sap) y
asientos producidos por motor_tiquipaya.construir_asiento() sobre fixtures
mínimos equivalentes a los de test_asiento.py. NO sube plantilla SAP real ni
datos contables reales al repositorio. sap_writer.py no repite ninguna
regla de ETAPA 5: estos tests alimentan directamente el asiento ya
construido y verifican que se serializa tal cual, sin recalcular nada.

Uso: python -m unittest tests.test_sap -v   (desde la raíz del repo)
"""

import datetime
import os
import sys
import shutil
import tempfile
import time
import unittest
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import motor_tiquipaya as motor
import sap_writer as sap
from tests import xlsx_fixtures as fx
from tests.xlsx_fixtures import crear_plantilla_sap
from tests.test_regresion_sintetica import (
    NOMBRE_CIERRE, _baseline_sfc101, _baseline_sfc102,
)


FECHA_CIERRE = "2026-08-19"


# ---------------------------------------------------------------------------
# Fixtures: asiento pequeño con los 4 orígenes de partida (HABER x2, VOUCHER,
# CI, ATC_NETO, ATC_COMISION), autocontenido y sin datos contables reales.
# ---------------------------------------------------------------------------

def _resultado_v2_ok(cuenta_ci="210201005"):
    detalle = {
        "sfc101_total": "1000.00",
        "sfc102_total": "500.00",
        "sfc101_haber": "1000.00",
        "sfc102_haber": "500.00",
        "alquileres_sfc101": "0.00",
        "alquileres_sfc102": "0.00",
        "vouchers_confirmados": [
            {
                "sfc": "SFC101",
                "importe": "300.00",
                "codigo_confirmado": "VCH0001",
                "codigo_informado": "VCH0001",
                "fecha_bancaria": "2026-08-20",  # nunca se usa para SAP; distinta a propósito
                "fecha_deposito": FECHA_CIERRE,
                "estado": "MATCH_EXACTO",
            },
        ],
        "ci_validas": [
            {
                "sfc": "SFC102",
                "referencia": "FAC-0001",  # ya no es fuente de texto_posicion
                "importe": "100.00",
                "cuenta_contable": cuenta_ci,
                "asignacion": "CI0001",
                "glosa": "PAGO PROVEEDOR AGUA",
                "fecha_ci": None,  # ver test_I: celda O debe quedar vacía
            },
        ],
        "atc_neto": {
            "importe": "1000.00",
            "codigo_confirmado": "ATC-19082026",
            "fecha_bancaria": FECHA_CIERRE,
        },
        "atc_comision": {"importe": "100.00"},
        "atc_aplica": True,
        "dolares": "0.00",
    }
    return {
        "fecha": FECHA_CIERRE,
        "universo_original": "1500.00",
        "alquileres": "0.00",
        "universo_ajustado": "1500.00",
        "componentes": {
            "vouchers": "300.00", "ci_operativas": "100.00",
            "atc_bruto": "1100.00", "dolares": "0.00",
        },
        "recaudacion_explicada": "1500.00",
        "diferencia": "0.00",
        "excepciones_bloqueantes": 0,
        "estado": "OK",
        "detalle": detalle,
    }


def _asiento_ok(cuenta_ci="210201005"):
    return motor.construir_asiento(_resultado_v2_ok(cuenta_ci))


def _metadata_cabecera():
    return {
        "tipo_asiento": "SA",
        "fecha_registro": "2026-08-19",
        "fecha_contabilizacion": "2026-08-19",
        "mes": "08",
        "texto_cabecera": "CAJA TIQUIPAYA 19-08-2026",
        "referencia": "TIQ-19082026",
    }


class _SapTestBase(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix="sap_test_")
        self.ruta_plantilla = os.path.join(self.tmpdir, "plantilla.xlsx")
        self.ruta_salida = os.path.join(self.tmpdir, "salida.xlsx")
        crear_plantilla_sap(self.ruta_plantilla)
        self.asiento = _asiento_ok()
        self.metadata = _metadata_cabecera()

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _generar(self):
        resultado = sap.generar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        self.assertEqual(resultado["estado"], "OK", resultado)
        return resultado

    def _reabrir(self):
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        return wb, wb["1"]


# ---------------------------------------------------------------------------
# A-M: generación básica y estructura de la plantilla
# ---------------------------------------------------------------------------

class TestGeneracionBasica(_SapTestBase):

    def test_A_asiento_valido_genera_archivo(self):
        self._generar()
        self.assertTrue(os.path.isfile(self.ruta_salida))

    def test_B_plantilla_original_no_se_modifica(self):
        hash_antes = sap._hash_archivo(self.ruta_plantilla)
        self._generar()
        hash_despues = sap._hash_archivo(self.ruta_plantilla)
        self.assertEqual(hash_antes, hash_despues)

    def test_C_hoja_1_obligatoria(self):
        ruta_mala = os.path.join(self.tmpdir, "plantilla_sin_hoja1.xlsx")
        crear_plantilla_sap(ruta_mala, hoja="Hoja1")
        resultado = sap.generar_sap(self.asiento, ruta_mala, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertIn("PLANTILLA_SIN_HOJA_1", resultado["problemas"])
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_D_cabecera_fila_10_correcta(self):
        self._generar()
        wb, ws = self._reabrir()
        self.assertEqual(ws["B10"].value, self.asiento["sociedad"])
        self.assertEqual(ws["C10"].value, self.metadata["tipo_asiento"])
        # ETAPA 8: D10/E10 son fecha Excel real (no texto); se comparan
        # normalizadas de vuelta a ISO 'YYYY-MM-DD'.
        self.assertEqual(sap._fecha_a_iso(ws["D10"].value), self.metadata["fecha_registro"])
        self.assertEqual(sap._fecha_a_iso(ws["E10"].value), self.metadata["fecha_contabilizacion"])
        self.assertEqual(ws["F10"].value, self.metadata["mes"])
        self.assertEqual(ws["G10"].value, self.metadata["texto_cabecera"])
        self.assertEqual(ws["H10"].value, "BOB")
        self.assertEqual(ws["L10"].value, self.metadata["referencia"])
        wb.close()

    def test_D2_cabecera_fechas_formato_corto_dd_mm_yyyy(self):
        # ETAPA 8: D10/E10 deben tener number_format compatible con
        # dd/mm/yyyy, y seguir siendo fechas válidas al reabrir.
        self._generar()
        wb, ws = self._reabrir()
        self.assertIn("dd/mm/yyyy", ws["D10"].number_format.lower())
        self.assertIn("dd/mm/yyyy", ws["E10"].number_format.lower())
        self.assertIsInstance(ws["D10"].value, datetime.datetime)
        self.assertIsInstance(ws["E10"].value, datetime.datetime)
        wb.close()

    def test_E_primera_partida_fila_16(self):
        self._generar()
        wb, ws = self._reabrir()
        primera = self.asiento["partidas"][0]
        self.assertEqual(ws["C16"].value, primera["cuenta_mayor"])
        self.assertEqual(ws["B16"].value, primera["sociedad"])
        wb.close()

    def test_F_multiples_partidas_mantienen_orden(self):
        self._generar()
        wb, ws = self._reabrir()
        cuentas_esperadas = [p["cuenta_mayor"] for p in self.asiento["partidas"]]
        cuentas_escritas = [
            ws[f"C{16 + i}"].value for i in range(len(cuentas_esperadas))
        ]
        self.assertEqual(cuentas_escritas, cuentas_esperadas)
        wb.close()

    def test_G_cargo_haber_escritos_correctamente(self):
        self._generar()
        wb, ws = self._reabrir()
        for i, p in enumerate(self.asiento["partidas"]):
            fila = 16 + i
            self.assertEqual(sap._decimal_celda(ws[f"E{fila}"].value), Decimal(p["cargo"]))
            self.assertEqual(sap._decimal_celda(ws[f"F{fila}"].value), Decimal(p["haber"]))
        wb.close()

    def test_H_fecha_valor_real_se_conserva(self):
        self._generar()
        wb, ws = self._reabrir()
        voucher = next(p for p in self.asiento["partidas"] if p["origen"] == "VOUCHER")
        idx = self.asiento["partidas"].index(voucher)
        self.assertEqual(sap._fecha_a_iso(ws[f"O{16 + idx}"].value), FECHA_CIERRE)
        self.assertIn("dd/mm/yyyy", ws[f"O{16 + idx}"].number_format.lower())
        wb.close()

    def test_I_fecha_valor_sin_fecha_propia_usa_fallback_fecha_cierre(self):
        # ETAPA 8: la CI de este fixture llega sin fecha_ci propia
        # (ver comentario en _resultado_v2_ok); ya no queda vacía, usa
        # el fallback autorizado (fecha del cierre) como fecha_valor.
        self._generar()
        wb, ws = self._reabrir()
        ci = next(p for p in self.asiento["partidas"] if p["origen"] == "CI")
        idx = self.asiento["partidas"].index(ci)
        self.assertEqual(ci["fecha_valor"], FECHA_CIERRE)
        self.assertEqual(sap._fecha_a_iso(ws[f"O{16 + idx}"].value), FECHA_CIERRE)
        wb.close()

    def test_J_asignacion_en_R(self):
        self._generar()
        wb, ws = self._reabrir()
        for i, p in enumerate(self.asiento["partidas"]):
            self.assertEqual(ws[f"R{16 + i}"].value, p.get("asignacion"))
        wb.close()

    def test_K_centro_beneficio_en_L(self):
        self._generar()
        wb, ws = self._reabrir()
        for i, p in enumerate(self.asiento["partidas"]):
            self.assertEqual(ws[f"L{16 + i}"].value, "10010101")
        wb.close()

    def test_L_total_cargo_total_haber_tras_reabrir(self):
        self._generar()
        wb, ws = self._reabrir()
        total_cargo = sum(
            (sap._decimal_celda(ws[f"E{16 + i}"].value) for i in range(len(self.asiento["partidas"]))),
            Decimal("0.00"),
        )
        total_haber = sum(
            (sap._decimal_celda(ws[f"F{16 + i}"].value) for i in range(len(self.asiento["partidas"]))),
            Decimal("0.00"),
        )
        self.assertEqual(total_cargo, Decimal(self.asiento["total_cargo"]))
        self.assertEqual(total_haber, Decimal(self.asiento["total_haber"]))
        self.assertEqual(total_cargo - total_haber, Decimal("0.00"))
        wb.close()

    def test_M_cantidad_de_partidas_coincide(self):
        self._generar()
        wb, ws = self._reabrir()
        self.assertEqual(sap._contar_filas_escritas(ws), len(self.asiento["partidas"]))
        wb.close()


# ---------------------------------------------------------------------------
# N-R: precondiciones — nunca se "arregla" nada, se bloquea
# ---------------------------------------------------------------------------

class TestPrecondicionesBloqueo(_SapTestBase):

    def test_N_asiento_error_no_genera_sap(self):
        resultado_v2 = _resultado_v2_ok()
        # Importe negativo en voucher bloquea ETAPA 5 (estado ERROR, partidas=[]).
        resultado_v2["detalle"]["vouchers_confirmados"][0]["importe"] = "-1.00"
        asiento_error = motor.construir_asiento(resultado_v2)
        self.assertEqual(asiento_error["estado"], "ERROR")

        resultado = sap.generar_sap(asiento_error, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_O1_usd_mayor_a_cero_genera_partida_en_sap(self):
        # CORRECCIÓN USD/DOLARES (post-ETAPA 8): ya no bloquea SAP; se
        # balancea el HABER SFC101 con el mismo importe para aislar esta
        # prueba del cuadre real de ejecutar_v2 (sin tocarlo).
        resultado_v2 = _resultado_v2_ok()
        resultado_v2["detalle"]["dolares"] = "50.00"
        resultado_v2["detalle"]["sfc101_haber"] = "1050.00"
        asiento_usd = motor.construir_asiento(resultado_v2)
        self.assertEqual(asiento_usd["estado"], "OK", asiento_usd)
        self.assertNotEqual(asiento_usd["estado"], "USD_CUENTA_PENDIENTE")

        resultado = sap.generar_sap(asiento_usd, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "OK", resultado)
        self.assertTrue(os.path.isfile(self.ruta_salida))

        usd = next(p for p in asiento_usd["partidas"] if p["origen"] == "DOLARES")
        idx = asiento_usd["partidas"].index(usd)
        fila = 16 + idx
        wb = openpyxl.load_workbook(self.ruta_salida, data_only=True)
        ws = wb["1"]
        self.assertEqual(ws[f"C{fila}"].value, "110101010")
        self.assertEqual(ws[f"D{fila}"].value, "RECAUDACION DOLARES")
        self.assertEqual(sap._decimal_celda(ws[f"E{fila}"].value), Decimal("50.00"))
        self.assertIsNone(ws[f"R{fila}"].value)
        wb.close()

    def test_O2_no_asiento_no_genera_sap(self):
        resultado_v2 = _resultado_v2_ok()
        resultado_v2["estado"] = "BLOQUEADO_EXCEPCION"
        resultado_v2["excepciones_bloqueantes"] = 1
        asiento_bloqueado = motor.construir_asiento(resultado_v2)
        self.assertEqual(asiento_bloqueado["estado"], "NO_ASIENTO")

        resultado = sap.generar_sap(asiento_bloqueado, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_P_diferencia_distinta_de_cero_no_genera_sap(self):
        asiento = _asiento_ok()
        asiento["diferencia"] = "0.01"
        resultado = sap.generar_sap(asiento, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertIn("DIFERENCIA_DISTINTA_DE_CERO", resultado["problemas"])
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_Q_metadata_cabecera_incompleta_error(self):
        metadata_incompleta = dict(self.metadata)
        del metadata_incompleta["referencia"]
        resultado = sap.generar_sap(self.asiento, self.ruta_plantilla, self.ruta_salida, metadata_incompleta)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertIn("METADATA_CABECERA_FALTANTE:referencia", resultado["problemas"])
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_R_ruta_salida_igual_a_plantilla_bloquea(self):
        resultado = sap.generar_sap(self.asiento, self.ruta_plantilla, self.ruta_plantilla, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertIn("RUTA_SALIDA_IGUAL_A_PLANTILLA", resultado["problemas"])


# ---------------------------------------------------------------------------
# S-U: defensa de cuentas ATC_COMISION / CI
# ---------------------------------------------------------------------------

class TestDefensaCuentas(_SapTestBase):

    def test_S_atc_comision_usa_110201008(self):
        atc_comision = next(p for p in self.asiento["partidas"] if p["origen"] == "ATC_COMISION")
        self.assertEqual(atc_comision["cuenta_mayor"], "110201008")
        self._generar()
        wb, ws = self._reabrir()
        idx = self.asiento["partidas"].index(atc_comision)
        self.assertEqual(ws[f"C{16 + idx}"].value, "110201008")
        wb.close()

    def test_T_atc_comision_110201003_bloquea(self):
        # ETAPA 5 ya bloquea esta cuenta para ATC_COMISION (estado ERROR),
        # por lo que aquí se construye un asiento manualmente para probar
        # la defensa PROPIA de ETAPA 6 en caso de recibir un asiento ya
        # armado fuera de construir_asiento().
        asiento_manipulado = {
            "fecha_cierre": FECHA_CIERRE,
            "sociedad": "BO01",
            "centro_beneficio": "10010101",
            "partidas": [
                {
                    "sociedad": "BO01", "cuenta_mayor": "110101001", "texto_posicion": None,
                    "cargo": "0.00", "haber": "100.00", "centro_beneficio": "10010101",
                    "fecha_valor": None, "asignacion": "SFC101", "origen": "UNIVERSO_SFC101",
                    "sfc_origen": "SFC101", "codigo_informado_original": None,
                },
                {
                    "sociedad": "BO01", "cuenta_mayor": "110201003", "texto_posicion": None,
                    "cargo": "100.00", "haber": "0.00", "centro_beneficio": "10010101",
                    "fecha_valor": None, "asignacion": "TIQUIPAYA AGO", "origen": "ATC_COMISION",
                    "sfc_origen": None, "codigo_informado_original": None,
                },
            ],
            "cantidad_partidas": 2,
            "total_cargo": "100.00", "total_haber": "100.00", "diferencia": "0.00",
            "correcciones_aplicadas": [], "estado": "OK", "problemas": [],
        }
        resultado = sap.generar_sap(asiento_manipulado, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertTrue(any(p.startswith("ATC_COMISION_CUENTA_PROHIBIDA") for p in resultado["problemas"]))
        self.assertFalse(os.path.isfile(self.ruta_salida))

    def test_U_ci_legitima_con_110201003_no_bloquea(self):
        asiento = _asiento_ok(cuenta_ci="110201003")
        self.assertEqual(asiento["estado"], "OK")
        resultado = sap.generar_sap(asiento, self.ruta_plantilla, self.ruta_salida, self.metadata)
        self.assertEqual(resultado["estado"], "OK")
        self.assertTrue(os.path.isfile(self.ruta_salida))


# ---------------------------------------------------------------------------
# V-W: orquestador y detección de alteración posterior
# ---------------------------------------------------------------------------

class TestValidacionPostEscritura(_SapTestBase):

    def test_V_plantilla_guardada_se_reabre_y_valida(self):
        resumen = sap.generar_y_validar_sap(
            self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata
        )
        self.assertEqual(resumen["estado_sap"], "OK", resumen)
        self.assertEqual(resumen["validacion"]["estado"], "OK")
        self.assertEqual(resumen["validacion"]["problemas"], [])
        self.assertEqual(resumen["validacion"]["diferencia"], "0.00")

    def test_W_archivo_manipulado_se_detecta(self):
        self._generar()

        wb = openpyxl.load_workbook(self.ruta_salida)
        ws = wb["1"]
        ws["E16"] = Decimal("999999.99")  # altera el CARGO de la primera partida
        wb.save(self.ruta_salida)
        wb.close()

        resultado = sap.validar_sap(
            self.ruta_salida, self.asiento, self.metadata, ruta_plantilla=self.ruta_plantilla
        )
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertTrue(any("CARGO_DISTINTO" in p or "DIFERENCIA" in p for p in resultado["problemas"]))

    def test_escritura_fuera_de_columna_autorizada_se_detecta(self):
        """Complementa W: si una columna NO autorizada de la fila de
        partidas es alterada (aquí "I", que en la plantilla trae "NO
        TOCAR"), validar_sap debe detectarlo aun si cargo/haber/cuenta
        siguen correctos."""
        self._generar()

        wb = openpyxl.load_workbook(self.ruta_salida)
        ws = wb["1"]
        ws["I16"] = "MANIPULADO"
        wb.save(self.ruta_salida)
        wb.close()

        resultado = sap.validar_sap(
            self.ruta_salida, self.asiento, self.metadata, ruta_plantilla=self.ruta_plantilla
        )
        self.assertEqual(resultado["estado"], "ERROR")
        self.assertIn("ESCRITURA_FUERA_DE_COLUMNA_AUTORIZADA:I16", resultado["problemas"])


# ---------------------------------------------------------------------------
# Rendimiento referencial (sección 16) — no es benchmark científico.
# ---------------------------------------------------------------------------

class TestRendimiento(_SapTestBase):

    def test_tiempos_referenciales_generacion_reapertura_validacion(self):
        t0 = time.perf_counter()
        generacion = sap.generar_sap(self.asiento, self.ruta_plantilla, self.ruta_salida, self.metadata)
        t1 = time.perf_counter()
        wb, ws = self._reabrir()
        wb.close()
        t2 = time.perf_counter()
        validacion = sap.validar_sap(
            self.ruta_salida, self.asiento, self.metadata, ruta_plantilla=self.ruta_plantilla
        )
        t3 = time.perf_counter()

        print(
            f"\n[rendimiento SAP] generación: {t1 - t0:.4f}s | "
            f"reapertura: {t2 - t1:.4f}s | validación: {t3 - t2:.4f}s"
        )

        self.assertEqual(generacion["estado"], "OK")
        self.assertEqual(validacion["estado"], "OK")


# ---------------------------------------------------------------------------
# ATC preconciliado (OPTIMIZACIÓN MAESTRO ÚNICO): asignación "REVISAR" ->
# se escribe literal en R y recibe relleno amarillo (solo visual).
# ---------------------------------------------------------------------------

def _asiento_ok_con_revisar():
    resultado_v2 = _resultado_v2_ok()
    resultado_v2["detalle"]["atc_neto"]["codigo_confirmado"] = "REVISAR"
    return motor.construir_asiento(resultado_v2)


class TestAtcRevisarEnSap(_SapTestBase):
    """Una partida ATC con asignación "REVISAR" (ver ETAPA 6 OPTIMIZACIÓN
    — ATC preconciliado) se escribe literal en columna R y recibe relleno
    amarillo. El amarillo es puramente visual: no afecta cargo/haber ni
    el cuadre del asiento."""

    def setUp(self):
        super().setUp()
        self.asiento = _asiento_ok_con_revisar()
        self.assertEqual(self.asiento["estado"], "OK")

    def test_revisar_se_escribe_literal_y_con_relleno_amarillo(self):
        self._generar()

        atc_neto = next(p for p in self.asiento["partidas"] if p["origen"] == "ATC_NETO")
        idx = self.asiento["partidas"].index(atc_neto)
        fila = 16 + idx
        self.assertEqual(atc_neto["asignacion"], "REVISAR")

        wb = openpyxl.load_workbook(self.ruta_salida)  # sin data_only: hace falta el estilo
        ws = wb["1"]
        celda = ws[f"R{fila}"]
        self.assertEqual(celda.value, "REVISAR")
        self.assertEqual(celda.fill.fill_type, "solid")
        self.assertEqual(celda.fill.fgColor.rgb, "00FFFF00")
        wb.close()

        validacion = sap.validar_sap(
            self.ruta_salida, self.asiento, self.metadata, ruta_plantilla=self.ruta_plantilla
        )
        self.assertEqual(validacion["estado"], "OK")
        self.assertEqual(validacion["diferencia"], "0.00")

    def test_asignacion_distinta_de_revisar_no_recibe_relleno(self):
        self._generar()
        atc_comision = next(p for p in self.asiento["partidas"] if p["origen"] == "ATC_COMISION")
        idx = self.asiento["partidas"].index(atc_comision)
        fila = 16 + idx

        wb = openpyxl.load_workbook(self.ruta_salida)
        ws = wb["1"]
        celda = ws[f"R{fila}"]
        self.assertNotEqual(celda.value, "REVISAR")
        self.assertIsNone(celda.fill.fill_type)
        wb.close()


# ---------------------------------------------------------------------------
# Integración completa: MAESTRO ÚNICO (MACROS + ATC TIQUIPAYA en un solo
# archivo) -> ejecutar_v2 -> construir_asiento -> generar_y_validar_sap.
# ---------------------------------------------------------------------------

class TestPipelineMaestroUnicoHastaSap(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix="sap_pipeline_")
        self.ruta_plantilla = os.path.join(self.tmpdir, "plantilla.xlsx")
        self.ruta_salida = os.path.join(self.tmpdir, "salida.xlsx")
        crear_plantilla_sap(self.ruta_plantilla)

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_pipeline_completo_ok(self):
        ruta_cierre = os.path.join(self.tmpdir, NOMBRE_CIERRE)
        ruta_maestro = os.path.join(self.tmpdir, "MACROS AGOSTO 2026.xlsm")
        fx.crear_cierre(ruta_cierre, _baseline_sfc101(), _baseline_sfc102())
        fx.crear_maestro_unico(
            ruta_maestro,
            macros_filas=[
                (FECHA_CIERRE, "VCH1001", "20000.00"),
                (FECHA_CIERRE, "VCH1002", "20805.00"),
                (FECHA_CIERRE, "VCH1003", "20000.00"),
                (FECHA_CIERRE, "VCH1004", "20000.00"),
            ],
            atc_filas=[
                (FECHA_CIERRE, "BANCO (NETO)", "110103012",
                 "ATC COCHABAMBA 19/08/2026", "130246.43", "3P02891953"),
                (FECHA_CIERRE, "COMISION ATC", "110201008",
                 "COMISION ATC 19/08/2026", "636.57", "TIQUIPAYA AGO"),
            ],
        )

        # Un solo archivo hace de MACROS y de ATC: la "segunda descarga"
        # queda eliminada (ruta_macros == ruta_atc).
        resultado_v2 = motor.ejecutar_v2(ruta_cierre, ruta_maestro, ruta_maestro)
        self.assertEqual(resultado_v2["estado"], "OK")
        asiento = motor.construir_asiento(resultado_v2)
        self.assertEqual(asiento["estado"], "OK")

        metadata = _metadata_cabecera()
        resumen = sap.generar_y_validar_sap(asiento, self.ruta_plantilla, self.ruta_salida, metadata)
        self.assertEqual(resumen["estado_sap"], "OK", resumen)
        self.assertEqual(resumen["validacion"]["problemas"], [])
        self.assertEqual(resumen["validacion"]["diferencia"], "0.00")


if __name__ == "__main__":
    unittest.main()
