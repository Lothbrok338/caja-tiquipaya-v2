"""
sap_writer.py — ETAPA 6: generación y validación determinística del archivo
SAP a partir del asiento YA VALIDADO por construir_asiento() (ETAPA 5).

ETAPA 6 no vuelve a decidir contabilidad. No recalcula cierres, no rehace
cruces, no busca MACROS ni ATC, no interpreta CI, no corrige importes, no
fuerza cuadre, no inventa cuentas ni fechas. El asiento de ETAPA 5 es la
única fuente de partidas: este módulo únicamente copia la plantilla SAP,
escribe las partidas ya construidas, guarda, reabre el archivo realmente
guardado y valida ese archivo contra el asiento fuente.

Estructura de la plantilla (hoja EXACTA "1"):

  CABECERA — fila 10:
    B=Sociedad C=TipoAsiento D=FechaRegistro E=FechaContabilizacion
    F=Mes G=TextoCabecera H=Moneda L=Referencia

  PARTIDAS — desde fila 16:
    B=Sociedad C=CuentaMayor D=TextoPosicion E=Cargo F=Haber
    L=CentroBeneficio O=FechaValor R=Asignacion U/V/W=XREF1/2/3

Las columnas intermedias (no listadas arriba) nunca se escriben y deben
permanecer intactas respecto de la plantilla original.

Uso:
    resumen = generar_y_validar_sap(asiento, ruta_plantilla, ruta_salida,
                                     metadata_cabecera)
"""

import hashlib
import os
import shutil
import time
import warnings
from decimal import Decimal

import openpyxl
from openpyxl.utils import get_column_letter

import excel_io as io


# ---------------------------------------------------------------------------
# Constantes de la plantilla (ver sección 3 de la especificación ETAPA 6)
# ---------------------------------------------------------------------------

_HOJA_SAP = "1"
_FILA_CABECERA = 10
_FILA_PRIMERA_PARTIDA = 16

_MONEDA = "BOB"

_COLUMNAS_AUTORIZADAS_CABECERA = {"B", "C", "D", "E", "F", "G", "H", "L"}
_COLUMNAS_AUTORIZADAS_PARTIDA = {"B", "C", "D", "E", "F", "L", "O", "R", "U", "V", "W"}

_CAMPOS_CABECERA_OBLIGATORIOS = (
    "tipo_asiento", "fecha_registro", "fecha_contabilizacion",
    "mes", "texto_cabecera", "referencia",
)

# 110201003 NO está prohibida globalmente: solo es inválida si la partida
# que la trae tiene origen ATC_COMISION (defensa propia de ETAPA 6, además
# de la que ya aplica ETAPA 5 en _validar_partidas). Una CI legítima puede
# usar esta cuenta sin problema.
_CUENTA_ATC_COMISION_PROHIBIDA = "110201003"


# ---------------------------------------------------------------------------
# Utilidades internas
# ---------------------------------------------------------------------------

def _abrir_libro(ruta, **kwargs):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return openpyxl.load_workbook(ruta, **kwargs)


def _hash_archivo(ruta):
    hasher = hashlib.sha256()
    with open(ruta, "rb") as f:
        for bloque in iter(lambda: f.read(1 << 16), b""):
            hasher.update(bloque)
    return hasher.hexdigest()


def _campo_faltante(valor):
    if valor is None:
        return True
    if isinstance(valor, str) and valor.strip() == "":
        return True
    return False


def _es_macro_habilitado(ruta):
    return str(ruta).lower().endswith((".xlsm", ".xltm"))


# ---------------------------------------------------------------------------
# Precondiciones (sección 12) — nunca se intenta "arreglar" nada aquí
# ---------------------------------------------------------------------------

def _validar_precondiciones(asiento, ruta_plantilla, ruta_salida, metadata_cabecera):
    problemas = []

    if asiento.get("estado") != "OK":
        problemas.append(f"ASIENTO_ESTADO_INVALIDO:{asiento.get('estado')}")
    if not asiento.get("partidas"):
        problemas.append("ASIENTO_SIN_PARTIDAS")

    total_cargo = asiento.get("total_cargo")
    total_haber = asiento.get("total_haber")
    diferencia = asiento.get("diferencia")
    if total_cargo is None or total_haber is None or diferencia is None:
        problemas.append("ASIENTO_SIN_TOTALES")
    else:
        if Decimal(total_cargo) != Decimal(total_haber):
            problemas.append("TOTAL_CARGO_DISTINTO_DE_TOTAL_HABER")
        if Decimal(diferencia) != 0:
            problemas.append("DIFERENCIA_DISTINTA_DE_CERO")

    if not ruta_plantilla or not os.path.isfile(ruta_plantilla):
        problemas.append("PLANTILLA_NO_ENCONTRADA")
    else:
        if os.path.abspath(str(ruta_salida)) == os.path.abspath(str(ruta_plantilla)):
            problemas.append("RUTA_SALIDA_IGUAL_A_PLANTILLA")
        try:
            wb_check = _abrir_libro(ruta_plantilla, read_only=True)
            hojas = list(wb_check.sheetnames)
            wb_check.close()
        except Exception as exc:  # noqa: BLE001 — plantilla ilegible, se reporta y se bloquea
            problemas.append(f"PLANTILLA_ILEGIBLE:{exc}")
            hojas = []
        if _HOJA_SAP not in hojas:
            problemas.append("PLANTILLA_SIN_HOJA_1")

    if not metadata_cabecera:
        problemas.append("METADATA_CABECERA_FALTANTE:TODOS")
    else:
        for campo in _CAMPOS_CABECERA_OBLIGATORIOS:
            if _campo_faltante(metadata_cabecera.get(campo)):
                problemas.append(f"METADATA_CABECERA_FALTANTE:{campo}")

    for i, p in enumerate(asiento.get("partidas") or []):
        if p.get("origen") == "ATC_COMISION" and p.get("cuenta_mayor") == _CUENTA_ATC_COMISION_PROHIBIDA:
            problemas.append(f"ATC_COMISION_CUENTA_PROHIBIDA:indice_{i}")

    return problemas


# ---------------------------------------------------------------------------
# Escritura (sección 6 y 7)
# ---------------------------------------------------------------------------

def _escribir_cabecera(ws, asiento, metadata_cabecera):
    fila = _FILA_CABECERA
    ws[f"B{fila}"] = asiento["sociedad"]
    ws[f"C{fila}"] = metadata_cabecera["tipo_asiento"]
    ws[f"D{fila}"] = metadata_cabecera["fecha_registro"]
    ws[f"E{fila}"] = metadata_cabecera["fecha_contabilizacion"]
    ws[f"F{fila}"] = metadata_cabecera["mes"]
    ws[f"G{fila}"] = metadata_cabecera["texto_cabecera"]
    ws[f"H{fila}"] = _MONEDA
    ws[f"L{fila}"] = metadata_cabecera["referencia"]


def _escribir_partidas(ws, partidas):
    for offset, p in enumerate(partidas):
        fila = _FILA_PRIMERA_PARTIDA + offset
        ws[f"B{fila}"] = p["sociedad"]
        ws[f"C{fila}"] = p["cuenta_mayor"]
        if p.get("texto_posicion") is not None:
            ws[f"D{fila}"] = p["texto_posicion"]

        celda_cargo = ws[f"E{fila}"]
        celda_haber = ws[f"F{fila}"]
        celda_cargo.value = Decimal(p["cargo"])
        celda_haber.value = Decimal(p["haber"])
        celda_cargo.number_format = "0.00"
        celda_haber.number_format = "0.00"

        ws[f"L{fila}"] = p["centro_beneficio"]
        if p.get("fecha_valor") is not None:
            ws[f"O{fila}"] = p["fecha_valor"]
        if p.get("asignacion") is not None:
            ws[f"R{fila}"] = p["asignacion"]
        if p.get("xref1") is not None:
            ws[f"U{fila}"] = p["xref1"]
        if p.get("xref2") is not None:
            ws[f"V{fila}"] = p["xref2"]
        if p.get("xref3") is not None:
            ws[f"W{fila}"] = p["xref3"]


def generar_sap(asiento, ruta_plantilla, ruta_salida, metadata_cabecera):
    """Copia `ruta_plantilla` a `ruta_salida` y escribe sobre la copia,
    exclusivamente, las partidas ya validadas de `asiento` (ETAPA 5).

    Nunca escribe sobre la plantilla original: si alguna precondición
    (sección 12) no se cumple, no copia ni escribe nada y devuelve
    estado ERROR con la lista de problemas encontrados.
    """
    problemas = _validar_precondiciones(asiento, ruta_plantilla, ruta_salida, metadata_cabecera)
    if problemas:
        return {"estado": "ERROR", "problemas": problemas, "ruta_salida": None}

    hash_plantilla_antes = _hash_archivo(ruta_plantilla)

    shutil.copyfile(ruta_plantilla, ruta_salida)

    keep_vba = _es_macro_habilitado(ruta_plantilla)
    wb = _abrir_libro(ruta_salida, keep_vba=keep_vba)
    try:
        ws = wb[_HOJA_SAP]
        _escribir_cabecera(ws, asiento, metadata_cabecera)
        _escribir_partidas(ws, asiento["partidas"])
        wb.save(ruta_salida)
    finally:
        wb.close()

    hash_plantilla_despues = _hash_archivo(ruta_plantilla)
    if hash_plantilla_antes != hash_plantilla_despues:
        # No debería poder ocurrir nunca (nunca se abre ruta_plantilla en
        # modo escritura): red de seguridad explícita, no un caso esperado.
        return {
            "estado": "ERROR",
            "problemas": ["PLANTILLA_ORIGINAL_MODIFICADA"],
            "ruta_salida": None,
        }

    return {
        "estado": "OK",
        "problemas": [],
        "ruta_salida": ruta_salida,
        "cantidad_partidas": len(asiento["partidas"]),
        "hash_plantilla": hash_plantilla_antes,
    }


# ---------------------------------------------------------------------------
# Validación post-escritura (sección 10) — siempre contra el archivo
# realmente guardado, reabierto desde disco.
# ---------------------------------------------------------------------------

def _validar_cabecera(ws, asiento, metadata_cabecera):
    problemas = []
    fila = _FILA_CABECERA
    esperado = {
        f"B{fila}": asiento.get("sociedad"),
        f"C{fila}": metadata_cabecera.get("tipo_asiento"),
        f"D{fila}": metadata_cabecera.get("fecha_registro"),
        f"E{fila}": metadata_cabecera.get("fecha_contabilizacion"),
        f"F{fila}": metadata_cabecera.get("mes"),
        f"G{fila}": metadata_cabecera.get("texto_cabecera"),
        f"H{fila}": _MONEDA,
        f"L{fila}": metadata_cabecera.get("referencia"),
    }
    for celda, valor_esperado in esperado.items():
        valor_real = ws[celda].value
        if valor_real != valor_esperado:
            problemas.append(
                f"CABECERA_{celda}_ESPERADO_{valor_esperado!r}_OBTENIDO_{valor_real!r}"
            )
    return problemas


def _contar_filas_escritas(ws):
    fila = _FILA_PRIMERA_PARTIDA
    cantidad = 0
    while ws[f"B{fila}"].value not in (None, "") or ws[f"C{fila}"].value not in (None, ""):
        cantidad += 1
        fila += 1
    return cantidad


def _decimal_celda(valor):
    if valor is None:
        return Decimal("0.00")
    return Decimal(io.money_str(valor))


def _validar_partidas_escritas(ws, partidas_esperadas):
    problemas = []

    cantidad_escrita = _contar_filas_escritas(ws)
    if cantidad_escrita != len(partidas_esperadas):
        problemas.append(
            "CANTIDAD_PARTIDAS_DISTINTA:"
            f"esperado_{len(partidas_esperadas)}_obtenido_{cantidad_escrita}"
        )

    total_cargo = Decimal("0.00")
    total_haber = Decimal("0.00")

    for offset, p in enumerate(partidas_esperadas):
        fila = _FILA_PRIMERA_PARTIDA + offset

        if ws[f"B{fila}"].value != p.get("sociedad"):
            problemas.append(f"PARTIDA_{fila}_SOCIEDAD_DISTINTA")
        if ws[f"C{fila}"].value != p.get("cuenta_mayor"):
            problemas.append(f"PARTIDA_{fila}_CUENTA_DISTINTA")
        if ws[f"D{fila}"].value != p.get("texto_posicion"):
            problemas.append(f"PARTIDA_{fila}_TEXTO_POSICION_DISTINTO")

        cargo_dec = _decimal_celda(ws[f"E{fila}"].value)
        haber_dec = _decimal_celda(ws[f"F{fila}"].value)
        if cargo_dec != Decimal(p.get("cargo", "0.00")):
            problemas.append(f"PARTIDA_{fila}_CARGO_DISTINTO")
        if haber_dec != Decimal(p.get("haber", "0.00")):
            problemas.append(f"PARTIDA_{fila}_HABER_DISTINTO")
        total_cargo += cargo_dec
        total_haber += haber_dec

        if ws[f"L{fila}"].value != p.get("centro_beneficio"):
            problemas.append(f"PARTIDA_{fila}_CENTRO_BENEFICIO_DISTINTO")
        if ws[f"O{fila}"].value != p.get("fecha_valor"):
            problemas.append(f"PARTIDA_{fila}_FECHA_VALOR_DISTINTA")
        if ws[f"R{fila}"].value != p.get("asignacion"):
            problemas.append(f"PARTIDA_{fila}_ASIGNACION_DISTINTA")

        for campo, col in (("xref1", "U"), ("xref2", "V"), ("xref3", "W")):
            if p.get(campo) is not None and ws[f"{col}{fila}"].value != p[campo]:
                problemas.append(f"PARTIDA_{fila}_{campo.upper()}_DISTINTO")

    return problemas, total_cargo, total_haber


def _validar_columnas_no_autorizadas(ws_sap, ws_plantilla, filas, columnas_autorizadas):
    problemas = []
    max_col = max(ws_sap.max_column or 1, ws_plantilla.max_column or 1, 30)
    for fila in filas:
        for idx in range(1, max_col + 1):
            col = get_column_letter(idx)
            if col in columnas_autorizadas:
                continue
            celda = f"{col}{fila}"
            if ws_sap[celda].value != ws_plantilla[celda].value:
                problemas.append(f"ESCRITURA_FUERA_DE_COLUMNA_AUTORIZADA:{celda}")
    return problemas


def validar_sap(ruta_sap, asiento, metadata_cabecera, ruta_plantilla=None):
    """Reabre `ruta_sap` desde disco (nunca confía en estructuras en
    memoria) y valida su contenido contra `asiento` (ETAPA 5) y
    `metadata_cabecera`. Si se pasa `ruta_plantilla`, también valida que
    ninguna celda fuera de las columnas autorizadas haya sido tocada.
    """
    if not os.path.isfile(ruta_sap):
        return {"estado": "ERROR", "problemas": ["ARCHIVO_SAP_NO_ENCONTRADO"]}

    if asiento.get("estado") != "OK" or not asiento.get("partidas"):
        return {"estado": "ERROR", "problemas": ["ASIENTO_NO_VALIDABLE"]}

    wb = _abrir_libro(ruta_sap, data_only=True)
    try:
        if _HOJA_SAP not in wb.sheetnames:
            return {"estado": "ERROR", "problemas": ["HOJA_1_NO_ENCONTRADA_EN_SAP"]}
        ws = wb[_HOJA_SAP]

        problemas = list(_validar_cabecera(ws, asiento, metadata_cabecera))

        partidas_esperadas = asiento["partidas"]
        problemas_partidas, total_cargo, total_haber = _validar_partidas_escritas(
            ws, partidas_esperadas
        )
        problemas.extend(problemas_partidas)

        diferencia = total_cargo - total_haber
        if diferencia != 0:
            problemas.append("DIFERENCIA_POST_ESCRITURA_DISTINTA_DE_CERO")

        if ruta_plantilla and os.path.isfile(ruta_plantilla):
            wb_plantilla = _abrir_libro(ruta_plantilla, data_only=True)
            try:
                if _HOJA_SAP in wb_plantilla.sheetnames:
                    ws_plantilla = wb_plantilla[_HOJA_SAP]
                    filas_partidas = list(range(
                        _FILA_PRIMERA_PARTIDA,
                        _FILA_PRIMERA_PARTIDA + len(partidas_esperadas),
                    ))
                    problemas.extend(_validar_columnas_no_autorizadas(
                        ws, ws_plantilla, [_FILA_CABECERA], _COLUMNAS_AUTORIZADAS_CABECERA
                    ))
                    problemas.extend(_validar_columnas_no_autorizadas(
                        ws, ws_plantilla, filas_partidas, _COLUMNAS_AUTORIZADAS_PARTIDA
                    ))
            finally:
                wb_plantilla.close()
    finally:
        wb.close()

    return {
        "estado": "OK" if not problemas else "ERROR",
        "problemas": problemas,
        "cantidad_partidas": len(partidas_esperadas),
        "total_cargo": io.money_str(total_cargo),
        "total_haber": io.money_str(total_haber),
        "diferencia": io.money_str(diferencia),
    }


# ---------------------------------------------------------------------------
# Orquestador (sección 11)
# ---------------------------------------------------------------------------

def generar_y_validar_sap(asiento, ruta_plantilla, ruta_salida, metadata_cabecera):
    """1) valida precondiciones, 2) copia plantilla, 3) escribe, 4) guarda,
    5) reabre, 6) valida, 7) devuelve un resumen estructurado."""
    t0 = time.perf_counter()
    generacion = generar_sap(asiento, ruta_plantilla, ruta_salida, metadata_cabecera)
    t1 = time.perf_counter()

    if generacion["estado"] != "OK":
        return {
            "estado_sap": "ERROR",
            "generacion": generacion,
            "validacion": None,
            "tiempo_generacion_s": t1 - t0,
            "tiempo_validacion_s": None,
        }

    validacion = validar_sap(ruta_salida, asiento, metadata_cabecera, ruta_plantilla=ruta_plantilla)
    t2 = time.perf_counter()

    return {
        "estado_sap": "OK" if validacion["estado"] == "OK" else "ERROR",
        "generacion": generacion,
        "validacion": validacion,
        "tiempo_generacion_s": t1 - t0,
        "tiempo_validacion_s": t2 - t1,
    }
