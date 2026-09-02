"""
excel_io.py — Capa de lectura determinística del CIERRE diario de Caja Tiquipaya (V2 CLOUD).

Responsabilidad única: leer un archivo CIERRE .xlsm (SOLO LECTURA) y devolver una
estructura Python normalizada con los datos crudos necesarios para el motor V2.

Esta etapa NO cruza contra MACROS, NO lee ATC mensual, NO cuadra, NO construye
asiento, NO genera SAP y NO usa Claude para procesar filas.

Tecnología: openpyxl (read_only=True, data_only=True) + Decimal para importes.
Nunca float para dinero.
"""

import os
import re
import unicodedata
import warnings
import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl


# ---------------------------------------------------------------------------
# Normalización de texto
# ---------------------------------------------------------------------------

def normalize_text(value):
    """Mayúsculas, sin tildes, espacios colapsados y sin espacios al borde."""
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_compact(value):
    """normalize_text sin espacios, para comparar nombres de hoja."""
    return normalize_text(value).replace(" ", "")


# ---------------------------------------------------------------------------
# Utilidades de importes (Decimal, nunca float)
# ---------------------------------------------------------------------------

def to_decimal(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal("0")
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # via str() para no arrastrar el error binario de los float
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        raise ValueError(f"No se pudo convertir a Decimal: {value!r}")


def money_str(value):
    dec = value if isinstance(value, Decimal) else to_decimal(value)
    return str(dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _texto_o_none(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


_RE_FECHA_ISO = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_RE_FECHA_DMY = re.compile(r"^(\d{2})[/-](\d{2})[/-](\d{4})$")


def _fecha_iso(value):
    """Normaliza un valor de fecha a 'YYYY-MM-DD'.

    Acepta datetime/date reales (lo habitual al leer una celda con formato
    de fecha en Excel). Si llega texto, solo se parsean formatos
    inequívocos ('YYYY-MM-DD' o 'DD-MM-YYYY'/'DD/MM/YYYY'); cualquier otro
    texto lanza ValueError en lugar de viajar sin validar hasta SAP.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()

    texto = str(value).strip()
    if not texto:
        return None

    m = _RE_FECHA_ISO.match(texto)
    if m:
        anio, mes, dia = m.groups()
        return datetime.date(int(anio), int(mes), int(dia)).isoformat()

    m = _RE_FECHA_DMY.match(texto)
    if m:
        dia, mes, anio = m.groups()
        return datetime.date(int(anio), int(mes), int(dia)).isoformat()

    raise ValueError(f"Fecha en formato no reconocido (no inequívoco): {value!r}")


# ---------------------------------------------------------------------------
# Localización de hojas
# ---------------------------------------------------------------------------

def _find_sfc_sheet(wb, suffix):
    objetivo = f"SFC{suffix}"
    for ws in wb.worksheets:
        if normalize_compact(ws.title) == objetivo:
            return ws
    raise ValueError(f"No se encontró la hoja {objetivo} en el archivo.")


def _find_ci_sheet(wb, suffix):
    for ws in wb.worksheets:
        compacto = normalize_compact(ws.title)
        if "COMUNICACIONESINTERNAS" in compacto and suffix in compacto:
            return ws
    raise ValueError(
        f"No se encontró la hoja de Comunicaciones Internas SFC{suffix} en el archivo."
    )


# ---------------------------------------------------------------------------
# Extracción: resumen SFC101 / SFC102 + composición de depósitos
# ---------------------------------------------------------------------------

_CAMPOS_RESUMEN = [
    "TOTAL MOVIMIENTO DEL DIA",
    "COBROS ATC",
    "TOTAL COMUNICACIONES INTERNAS",
    "DOLARES",
]


def _leer_resumen_sfc(ws, sfc_label):
    rows = list(ws.iter_rows(values_only=True))

    campos = {k: None for k in _CAMPOS_RESUMEN}
    for row in rows:
        if not row:
            continue
        etiqueta = normalize_text(row[0])
        if etiqueta in campos and campos[etiqueta] is None:
            valor = row[1] if len(row) > 1 else None
            campos[etiqueta] = to_decimal(valor)

    faltantes = [k for k, v in campos.items() if v is None]
    if faltantes:
        raise ValueError(f"{sfc_label}: no se encontraron los campos: {faltantes}")

    depositos = _leer_composicion_depositos(rows, sfc_label)

    return {
        "total_movimiento": money_str(campos["TOTAL MOVIMIENTO DEL DIA"]),
        "cobros_atc": money_str(campos["COBROS ATC"]),
        "total_ci": money_str(campos["TOTAL COMUNICACIONES INTERNAS"]),
        "dolares": money_str(campos["DOLARES"]),
        "depositos": depositos,
    }


def _leer_composicion_depositos(rows, sfc_label):
    header_row_idx = None
    comp_idx = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            text = normalize_text(cell)
            if "COMPOSICION" in text and "DEPOSITO" in text:
                header_row_idx = i
                comp_idx = j
                break
        if comp_idx is not None:
            break

    if comp_idx is None:
        raise ValueError(f"{sfc_label}: no se encontró la sección COMPOSICIÓN DE DEPÓSITOS.")

    header_row = rows[header_row_idx]
    importe_idx = fecha_idx = asignacion_idx = banco_idx = None
    for j in range(comp_idx + 1, len(header_row)):
        text = normalize_text(header_row[j])
        if not text:
            continue
        if importe_idx is None and "IMPORTE" in text:
            importe_idx = j
        elif fecha_idx is None and "FECHA" in text and "DEPOSITO" in text:
            fecha_idx = j
        elif asignacion_idx is None and "ASIGNACION" in text:
            asignacion_idx = j
        elif banco_idx is None and "BANCO" in text:
            banco_idx = j

    faltantes = []
    if importe_idx is None:
        faltantes.append("IMPORTE Bs")
    if fecha_idx is None:
        faltantes.append("FECHA DE DEPOSITO")
    if asignacion_idx is None:
        faltantes.append("ASIGNACION")
    if banco_idx is None:
        faltantes.append("BANCO")
    if faltantes:
        raise ValueError(
            f"{sfc_label}: faltan columnas en COMPOSICIÓN DE DEPÓSITOS: {faltantes}"
        )

    depositos = []
    for row in rows[header_row_idx + 1:]:
        comp_val = row[comp_idx] if comp_idx < len(row) else None
        comp_text = normalize_text(comp_val)
        if not comp_text:
            # Fila vacía/separadora dentro del bloque: no trunca, se ignora
            # y se sigue leyendo (puede haber depósitos válidos después).
            continue
        if "DEPOSITO" not in comp_text:
            # Contenido real que no es un depósito: fin determinístico del
            # bloque (estructura fija según sección 4). No se lee nada más
            # allá de este punto.
            break

        importe_val = row[importe_idx] if importe_idx < len(row) else None
        importe_dec = to_decimal(importe_val)
        if importe_val is None or importe_dec <= 0:
            continue  # solo filas DEPOSITO con importe válido > 0

        fecha_val = row[fecha_idx] if fecha_idx < len(row) else None
        asignacion_val = row[asignacion_idx] if asignacion_idx < len(row) else None
        banco_val = row[banco_idx] if banco_idx < len(row) else None

        depositos.append({
            "sfc": sfc_label,
            "deposito": _texto_o_none(comp_val) or "",
            "importe": money_str(importe_dec),
            "fecha_deposito": _fecha_iso(fecha_val),
            "asignacion": _texto_o_none(asignacion_val),
            "banco_cuenta": _texto_o_none(banco_val),
        })

    return depositos


# ---------------------------------------------------------------------------
# Extracción: Comunicaciones Internas
# ---------------------------------------------------------------------------

def _leer_comunicaciones_internas(ws, sfc_label):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    _ENCABEZADOS_FECHA_CI = {"FECHA", "FECHA CI", "FECHA COMUNICACION INTERNA"}
    idx_n = idx_factura = idx_total = idx_cuenta = idx_asignacion = idx_banco = None
    idx_fecha = None
    for j, cell in enumerate(header):
        text = normalize_text(cell)
        if not text:
            continue
        compacto = text.replace("°", "").replace("º", "").strip()
        if idx_n is None and compacto == "N":
            idx_n = j
        elif idx_factura is None and "FACTURA" in text:
            idx_factura = j
        elif idx_total is None and "TOTAL C.I" in text:
            idx_total = j
        elif idx_cuenta is None and "CUENTA CONTABLE" in text:
            idx_cuenta = j
        elif idx_asignacion is None and "ASIGNACION" in text:
            idx_asignacion = j
        elif idx_banco is None and text == "BANCO":
            idx_banco = j
        elif idx_fecha is None and text in _ENCABEZADOS_FECHA_CI:
            idx_fecha = j

    faltantes = []
    if idx_factura is None:
        faltantes.append("NUMERO DE FACTURA")
    if idx_total is None:
        faltantes.append("TOTAL C.I.")
    if idx_cuenta is None:
        faltantes.append("CUENTA CONTABLE BANCO")
    if idx_asignacion is None:
        faltantes.append("ASIGNACION")
    if idx_banco is None:
        faltantes.append("BANCO")
    if faltantes:
        raise ValueError(
            f"{sfc_label}: faltan columnas en Comunicaciones Internas: {faltantes}"
        )

    idx_fila_valida = idx_n if idx_n is not None else idx_factura

    resultado = []
    for row in rows[1:]:
        marcador = row[idx_fila_valida] if idx_fila_valida < len(row) else None
        if marcador is None:
            continue  # fila en blanco / fila de total, no es una CI real

        referencia = row[idx_factura] if idx_factura < len(row) else None
        importe_val = row[idx_total] if idx_total < len(row) else None
        cuenta_val = row[idx_cuenta] if idx_cuenta < len(row) else None
        asignacion_val = row[idx_asignacion] if idx_asignacion < len(row) else None
        banco_val = row[idx_banco] if idx_banco < len(row) else None
        fecha_val = None
        if idx_fecha is not None and idx_fecha < len(row):
            fecha_val = row[idx_fecha]

        banco_texto = _texto_o_none(banco_val)
        es_alquileres = banco_texto is not None and normalize_text(banco_texto) == "ALQUILERES"

        resultado.append({
            "sfc": sfc_label,
            "referencia": _texto_o_none(referencia),
            "importe": money_str(to_decimal(importe_val)),
            "cuenta_contable": _texto_o_none(cuenta_val),
            "asignacion": _texto_o_none(asignacion_val),
            "banco": banco_texto,
            "alquileres": es_alquileres,
            # Fecha propia de la CI si la hoja la trae; nunca se rellena
            # con la fecha del cierre (ver excel_io._fecha_iso).
            "fecha_ci": _fecha_iso(fecha_val) if fecha_val is not None else None,
        })

    return resultado


# ---------------------------------------------------------------------------
# Fecha de cierre (desde el nombre del archivo, NUNCA desde config regional)
# ---------------------------------------------------------------------------

def _extraer_fecha_de_nombre(ruta_archivo):
    nombre = os.path.basename(str(ruta_archivo))
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", nombre)
    if not m:
        raise ValueError(
            f"No se pudo interpretar la fecha del cierre desde el nombre de archivo: {nombre}"
        )
    dia, mes, anio = m.groups()
    return datetime.date(int(anio), int(mes), int(dia)).isoformat()


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------

def leer_cierre(ruta_archivo):
    """
    Lee un archivo CIERRE .xlsm UNA sola vez (read_only, data_only) y devuelve:

    {
      "fecha_cierre": "YYYY-MM-DD",
      "sfc101": {
        "total_movimiento": "...", "cobros_atc": "...", "total_ci": "...",
        "dolares": "...", "depositos": [...]
      },
      "sfc102": { ... igual estructura ... },
      "comunicaciones_internas": [...]
    }

    Todos los importes se serializan como strings con 2 decimales.
    No cruza contra MACROS ni ATC. No cuadra. No modifica el archivo fuente.
    """
    fecha_cierre = _extraer_fecha_de_nombre(ruta_archivo)

    # openpyxl en read_only difiere el parseo de cada hoja hasta que se itera
    # sobre ella, así que el warning de "Data Validation extension" puede
    # aparecer tanto en load_workbook() como al iterar filas más abajo.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(
            ruta_archivo, read_only=True, data_only=True, keep_vba=False
        )
        try:
            ws_sfc101 = _find_sfc_sheet(wb, "101")
            ws_sfc102 = _find_sfc_sheet(wb, "102")
            ws_ci101 = _find_ci_sheet(wb, "101")
            ws_ci102 = _find_ci_sheet(wb, "102")

            sfc101 = _leer_resumen_sfc(ws_sfc101, "SFC101")
            sfc102 = _leer_resumen_sfc(ws_sfc102, "SFC102")
            ci101 = _leer_comunicaciones_internas(ws_ci101, "SFC101")
            ci102 = _leer_comunicaciones_internas(ws_ci102, "SFC102")
        finally:
            wb.close()

    return {
        "fecha_cierre": fecha_cierre,
        "sfc101": sfc101,
        "sfc102": sfc102,
        "comunicaciones_internas": ci101 + ci102,
    }


# ---------------------------------------------------------------------------
# MACROS BNB (mensual) — vouchers y NETO ATC únicamente
# ---------------------------------------------------------------------------
#
# Fuente: hoja EXACTA "Tablas Dinamicas Profesional" del archivo MACROS
# mensual. NO se lee ninguna otra pestaña. Columnas relevantes:
#   Fecha, Código de Asignación, Créditos
#
# Esta hoja trae, dentro del propio rango de datos, filas de encabezado
# repetidas (el mismo rótulo de columna vuelto a insertar cada cierto
# bloque). Se descartan de forma determinística por contenido/estructura
# (comparando cada fila contra el encabezado real leído en esta misma
# pasada), sin depender de una cantidad fija de repeticiones.
#
# Se abre UNA sola vez y se construyen dos índices en memoria:
#   por_codigo:  {codigo_normalizado: [movimiento, ...]}
#   por_importe: {"importe_2dec": [movimiento, ...]}
#
# NO se usa para CI. NO se exploran otros movimientos fuera de estos índices.

_MACROS_HOJA_BNB = "Tablas Dinamicas Profesional"


def normalize_codigo(value):
    """Mayúsculas, sin espacios internos ni al borde. Para comparar códigos de asignación."""
    if value is None:
        return ""
    return normalize_text(value).replace(" ", "")


def leer_macros_bnb(ruta_archivo, hoja=_MACROS_HOJA_BNB):
    """
    Abre MACROS mensual UNA sola vez (read_only, data_only) y devuelve índices
    en memoria sobre la hoja de movimientos BNB crudos:

    {
      "por_codigo": {codigo_normalizado: [movimiento, ...]},
      "por_importe": {"importe_2dec_str": [movimiento, ...]},
    }

    Cada movimiento:
      {"codigo": "...", "importe": "0.00", "fecha": "YYYY-MM-DD"}

    No recorre la hoja completa por cada búsqueda posterior: los índices ya
    quedan armados en memoria tras esta única pasada.
    """
    por_codigo = {}
    por_importe = {}

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(
            ruta_archivo, read_only=True, data_only=True, keep_vba=False
        )
        try:
            if hoja not in wb.sheetnames:
                raise ValueError(f"No se encontró la hoja '{hoja}' en MACROS.")
            ws = wb[hoja]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                raise ValueError(f"MACROS: la hoja '{hoja}' está vacía.")

            idx = {normalize_text(h): j for j, h in enumerate(header) if h is not None}
            idx_codigo = idx.get("CODIGO DE ASIGNACION")
            idx_creditos = idx.get("CREDITOS")
            idx_fecha = idx.get("FECHA")
            faltantes = []
            if idx_codigo is None:
                faltantes.append("CODIGO DE ASIGNACION")
            if idx_creditos is None:
                faltantes.append("Créditos")
            if idx_fecha is None:
                faltantes.append("Fecha")
            if faltantes:
                raise ValueError(f"MACROS: faltan columnas en '{hoja}': {faltantes}")

            # Texto normalizado del encabezado real, para detectar filas de
            # encabezado repetidas por contenido (no por posición ni conteo).
            header_fecha_texto = normalize_text(header[idx_fecha]) if idx_fecha < len(header) else ""
            header_codigo_texto = normalize_text(header[idx_codigo]) if idx_codigo < len(header) else ""

            for row in rows:
                codigo_val = row[idx_codigo] if idx_codigo < len(row) else None
                credito_val = row[idx_creditos] if idx_creditos < len(row) else None
                if codigo_val is None or credito_val is None:
                    continue  # solo movimientos de crédito con código (depósitos/abonos)

                fecha_val_bruta = row[idx_fecha] if idx_fecha < len(row) else None
                if (
                    normalize_text(fecha_val_bruta) == header_fecha_texto
                    or normalize_text(codigo_val) == header_codigo_texto
                ):
                    continue  # fila de encabezado repetida dentro del rango de datos

                codigo_norm = normalize_codigo(codigo_val)
                if not codigo_norm:
                    continue
                importe_dec = to_decimal(credito_val)
                if importe_dec <= 0:
                    continue

                movimiento = {
                    "codigo": _texto_o_none(codigo_val),
                    "importe": money_str(importe_dec),
                    "fecha": _fecha_iso(fecha_val_bruta),
                }

                por_codigo.setdefault(codigo_norm, []).append(movimiento)
                por_importe.setdefault(movimiento["importe"], []).append(movimiento)
        finally:
            wb.close()

    return {"por_codigo": por_codigo, "por_importe": por_importe}


# ---------------------------------------------------------------------------
# ATC mensual — NETO y COMISIÓN por fecha
# ---------------------------------------------------------------------------
#
# Dos formatos posibles, detectados por el NOMBRE de hoja (nunca por un
# parámetro que el llamador deba adivinar):
#
# LEGADO (comportamiento histórico, sin cambios): archivo ATC mensual
#   separado, hoja única (la primera del workbook) con columnas FECHA,
#   TIPO, MONTO (CUENTA CONTABLE/DETALLE/ASIGNACION pueden estar
#   presentes pero no se usan). Activo cuando el archivo NO trae una
#   hoja llamada exactamente "ATC TIQUIPAYA". Este ATC se cruza después
#   contra MACROS (ver motor_tiquipaya.cruzar_atc): NO trae cuenta ni
#   asignación resueltas.
#
# PRECONCILIADO (ETAPA 6 — maestro único): activo cuando el archivo SÍ
#   trae una hoja "ATC TIQUIPAYA" (puede ser el mismo archivo mensual de
#   MACROS: "MAESTRO MENSUAL"). El ATC ya viene conciliado con columnas
#   FECHA, TIPO, CUENTA CONTABLE, DETALLE, MONTO, ASIGNACION: NUNCA se
#   vuelve a cruzar contra MACROS (ver motor_tiquipaya.cruzar_atc_preconciliado).
#
# TIPO ∈ {"BANCO (NETO)", "COMISIÓN ATC"} en ambos formatos.

_ATC_HOJA_PRECONCILIADA = "ATC TIQUIPAYA"


def leer_atc_mensual(ruta_archivo):
    """
    Abre el archivo ATC UNA sola vez (read_only, data_only) y devuelve:

    {
      "modo": "LEGADO" | "PRECONCILIADO",
      "por_fecha": { "YYYY-MM-DD": {...} },
    }

    LEGADO — por_fecha:
      { "YYYY-MM-DD": {"neto": "0.00" | None, "comision": "0.00" | None} }

    PRECONCILIADO — por_fecha:
      { "YYYY-MM-DD": {
          "neto": {"monto": "0.00", "cuenta_contable": "...",
                    "detalle": "...", "asignacion": "..."} | None,
          "comision": {...misma forma...} | None,
      } }

    La ASIGNACION de PRECONCILIADO se conserva literal (incluido el
    valor "REVISAR"): esta capa de lectura no la interpreta ni la
    reemplaza.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(
            ruta_archivo, read_only=True, data_only=True, keep_vba=False
        )
        try:
            ws_preconciliada = _find_atc_preconciliado_sheet(wb)
            if ws_preconciliada is not None:
                por_fecha = _leer_atc_preconciliado(ws_preconciliada)
                modo = "PRECONCILIADO"
            else:
                por_fecha = _leer_atc_legado(wb[wb.sheetnames[0]])
                modo = "LEGADO"
        finally:
            wb.close()

    return {"modo": modo, "por_fecha": por_fecha}


def _find_atc_preconciliado_sheet(wb):
    """Busca la hoja "ATC TIQUIPAYA" por nombre normalizado (mayúsculas,
    sin tildes, espacios colapsados), igual que _find_sfc_sheet/
    _find_ci_sheet, para no depender de mayúsculas/espacios exactos.
    None si el workbook no la trae (flujo LEGADO)."""
    objetivo = normalize_compact(_ATC_HOJA_PRECONCILIADA)
    for ws in wb.worksheets:
        if normalize_compact(ws.title) == objetivo:
            return ws
    return None


def _leer_atc_legado(ws):
    """Comportamiento histórico, sin ningún cambio de reglas."""
    por_fecha = {}

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError("ATC mensual: hoja vacía.")

    idx = {normalize_text(h): j for j, h in enumerate(header) if h is not None}
    idx_fecha = idx.get("FECHA")
    idx_tipo = idx.get("TIPO")
    idx_monto = idx.get("MONTO")
    faltantes = []
    if idx_fecha is None:
        faltantes.append("FECHA")
    if idx_tipo is None:
        faltantes.append("TIPO")
    if idx_monto is None:
        faltantes.append("MONTO")
    if faltantes:
        raise ValueError(f"ATC mensual: faltan columnas: {faltantes}")

    for row in rows:
        fecha_val = row[idx_fecha] if idx_fecha < len(row) else None
        tipo_val = row[idx_tipo] if idx_tipo < len(row) else None
        monto_val = row[idx_monto] if idx_monto < len(row) else None
        if fecha_val is None or tipo_val is None or monto_val is None:
            continue

        fecha_iso = _fecha_iso(fecha_val)
        tipo_norm = normalize_text(tipo_val)
        importe = money_str(to_decimal(monto_val))

        registro = por_fecha.setdefault(fecha_iso, {"neto": None, "comision": None})
        if "NETO" in tipo_norm:
            if registro["neto"] is not None:
                # Fila NETO duplicada para la misma fecha: no hay
                # regla inequívoca para consolidarlas (no se suma,
                # no se usa "la última fila"). Falla cerrado.
                raise ValueError(
                    f"ATC mensual: fila NETO duplicada para fecha {fecha_iso}."
                )
            registro["neto"] = importe
        elif "COMISION" in tipo_norm:
            if registro["comision"] is not None:
                raise ValueError(
                    f"ATC mensual: fila COMISIÓN duplicada para fecha {fecha_iso}."
                )
            registro["comision"] = importe

    return por_fecha


def _leer_atc_preconciliado(ws):
    """Hoja "ATC TIQUIPAYA": ATC ya conciliado. Cuenta contable, detalle,
    monto y asignación se toman literalmente de la hoja, fila por fila,
    sin cruzar contra MACROS ni interpretar la asignación."""
    por_fecha = {}

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError("ATC TIQUIPAYA: hoja vacía.")

    idx = {normalize_text(h): j for j, h in enumerate(header) if h is not None}
    idx_fecha = idx.get("FECHA")
    idx_tipo = idx.get("TIPO")
    idx_cuenta = idx.get("CUENTA CONTABLE")
    idx_detalle = idx.get("DETALLE")
    idx_monto = idx.get("MONTO")
    idx_asignacion = idx.get("ASIGNACION")
    faltantes = []
    if idx_fecha is None:
        faltantes.append("FECHA")
    if idx_tipo is None:
        faltantes.append("TIPO")
    if idx_cuenta is None:
        faltantes.append("CUENTA CONTABLE")
    if idx_detalle is None:
        faltantes.append("DETALLE")
    if idx_monto is None:
        faltantes.append("MONTO")
    if idx_asignacion is None:
        faltantes.append("ASIGNACION")
    if faltantes:
        raise ValueError(f"ATC TIQUIPAYA: faltan columnas: {faltantes}")

    for row in rows:
        fecha_val = row[idx_fecha] if idx_fecha < len(row) else None
        tipo_val = row[idx_tipo] if idx_tipo < len(row) else None
        monto_val = row[idx_monto] if idx_monto < len(row) else None
        if fecha_val is None or tipo_val is None or monto_val is None:
            continue

        fecha_iso = _fecha_iso(fecha_val)
        tipo_norm = normalize_text(tipo_val)
        cuenta_val = row[idx_cuenta] if idx_cuenta < len(row) else None
        detalle_val = row[idx_detalle] if idx_detalle < len(row) else None
        asignacion_val = row[idx_asignacion] if idx_asignacion < len(row) else None

        entrada = {
            "monto": money_str(to_decimal(monto_val)),
            "cuenta_contable": _texto_o_none(cuenta_val),
            "detalle": _texto_o_none(detalle_val),
            # Literal, incluido "REVISAR": no se interpreta ni se
            # reemplaza en esta capa de lectura.
            "asignacion": _texto_o_none(asignacion_val),
        }

        registro = por_fecha.setdefault(fecha_iso, {"neto": None, "comision": None})
        if "NETO" in tipo_norm:
            if registro["neto"] is not None:
                raise ValueError(
                    f"ATC TIQUIPAYA: fila NETO duplicada para fecha {fecha_iso}."
                )
            registro["neto"] = entrada
        elif "COMISION" in tipo_norm:
            if registro["comision"] is not None:
                raise ValueError(
                    f"ATC TIQUIPAYA: fila COMISIÓN duplicada para fecha {fecha_iso}."
                )
            registro["comision"] = entrada

    return por_fecha


# ---------------------------------------------------------------------------
# Prueba de regresión manual (no forma parte del motor)
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import json

    ruta = sys.argv[1] if len(sys.argv) > 1 else None
    if not ruta:
        print("Uso: python excel_io.py <ruta_al_cierre.xlsm>")
        raise SystemExit(1)

    data = leer_cierre(ruta)
    print(json.dumps(data, ensure_ascii=False, indent=2))
